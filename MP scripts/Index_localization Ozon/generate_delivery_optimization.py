#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор отчёта по оптимизации среднего времени доставки OZON (v2)

Анализирует время доставки, остатки по складам, валидные маршруты перемещений
и статусы артикулов. Генерирует рекомендации только по реальному спросу
и только для маршрутов, доступных в OZON.

Запуск:
    python3 generate_delivery_optimization.py "data ИП Ozon/10.02"
    python3 generate_delivery_optimization.py "data ИП Ozon/10.02" --safety-days 7
    python3 generate_delivery_optimization.py "data ИП Ozon/10.02" --no-statuses
"""

import argparse
import math
import re
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Путь к корню проекта для импорта data_layer
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from scripts.data_layer import get_artikuly_statuses
    HAS_DATA_LAYER = True
except ImportError:
    HAS_DATA_LAYER = False


# ============================================
# КОНФИГУРАЦИЯ
# ============================================

BASE_PATH = Path(__file__).parent

# Пороги времени доставки (часы)
TARGET_DELIVERY_TIME = 29
CRITICAL_TIME = 90
ATTENTION_TIME = 60
HIGH_VOLUME = 3

# Горизонт планирования (недели)
DEFAULT_PLANNING_WEEKS = 8

# Защита доноров
DEFAULT_SAFETY_DAYS = 14

# Статусы ликвидности: из каких кластеров НЕЛЬЗЯ забирать товар
BLOCKED_STATUSES = {
    'Дефицитный', 'Был дефицитный',
    'Очень популярный',
    'Ожидаем поставки',
}

CAUTIOUS_STATUSES = {
    'Популярный', 'Был популярный',
    'Был очень популярный',
}

# Международные кластеры (перемещение невозможно)
INTERNATIONAL_CLUSTERS = {
    'Астана', 'Алматы', 'Беларусь',
    'Армения', 'Азербайджан', 'Грузия',
    'Кыргызстан', 'Узбекистан',
    'Калининград',
}

# Наценка за нелокальную продажу (% от цены товара)
SURCHARGE_RATES = {
    'Москва, МО и Дальние регионы': 0.08,
    'Санкт-Петербург и СЗО': 0.08,
    'Краснодар': 0.06, 'Красноярск': 0.06,
    'Махачкала': 0.06, 'Невинномысск': 0.06,
    'Омск': 0.06, 'Оренбург': 0.06,
    'Пермь': 0.06, 'Самара': 0.06,
    'Саратов': 0.06, 'Тверь': 0.06,
    'Ярославль': 0.06,
    'Воронеж': 0.04, 'Екатеринбург': 0.04,
    'Тюмень': 0.04, 'Уфа': 0.04,
    'Дальний Восток': 0.0, 'Казань': 0.0,
    'Калининград': 0.0, 'Новосибирск': 0.0,
    'Ростов': 0.0,
}

DEFAULT_MOVEMENT_COST = 38
DEFAULT_AVG_PRICE = 2000
OZON_MIN_MOVEMENT = 200
OZON_MAX_MOVEMENT = 10000

# Маппинг склад-назначение → кластер (из справочника перемещений)
DEST_WAREHOUSE_TO_CLUSTER = {
    'ВОРОНЕЖ_2_РФЦ': 'Воронеж',
    'ЕКАТЕРИНБУРГ_РФЦ_НОВЫЙ': 'Екатеринбург',
    'КАЗАНЬ_РФЦ_НОВЫЙ': 'Казань',
    'КРАСНОЯРСК_МРФЦ': 'Красноярск',
    'НЕВИННОМЫССК_РФЦ': 'Невинномысск',
    'ОМСК_РФЦ': 'Омск',
    'ОРЕНБУРГ_РФЦ': 'Оренбург',
    'ПУШКИНО_2_РФЦ': 'Москва, МО и Дальние регионы',
    'РОСТОВ-НА-ДОНУ_РФЦ': 'Ростов',
    'САМАРА_РФЦ': 'Самара',
    'СПБ_БУГРЫ_РФЦ': 'Санкт-Петербург и СЗО',
    'САРАТОВ_РФЦ': 'Саратов',
    'ТЮМЕНЬ_РФЦ': 'Тюмень',
    'ХАБАРОВСК_2_РФЦ': 'Дальний Восток',
    'ЯРОСЛАВЛЬ_РФЦ': 'Ярославль',
}


# ============================================
# CLI
# ============================================

def parse_args():
    parser = argparse.ArgumentParser(
        description='Отчёт по оптимизации среднего времени доставки OZON (v2)'
    )
    parser.add_argument(
        'data_folder',
        help='Папка с данными, например "data ИП Ozon/10.02"'
    )
    parser.add_argument(
        '--safety-days', type=int, default=DEFAULT_SAFETY_DAYS,
        help=f'Дней запаса для донора (по умолчанию {DEFAULT_SAFETY_DAYS})'
    )
    parser.add_argument(
        '--target-time', type=int, default=TARGET_DELIVERY_TIME,
        help=f'Целевое время доставки, часы (по умолчанию {TARGET_DELIVERY_TIME})'
    )
    parser.add_argument(
        '--planning-weeks', type=int, default=DEFAULT_PLANNING_WEEKS,
        help=f'Горизонт планирования, недели (по умолчанию {DEFAULT_PLANNING_WEEKS})'
    )
    parser.add_argument(
        '--movement-cost-per-unit', type=float, default=DEFAULT_MOVEMENT_COST,
        help=f'Стоимость перемещения 1 единицы, руб (по умолчанию {DEFAULT_MOVEMENT_COST})'
    )
    parser.add_argument(
        '--avg-product-price', type=float, default=DEFAULT_AVG_PRICE,
        help=f'Средняя цена товара, руб (по умолчанию {DEFAULT_AVG_PRICE})'
    )
    parser.add_argument(
        '--no-statuses', action='store_true', default=False,
        help='Не загружать статусы из Supabase'
    )
    parser.add_argument(
        '--template',
        default=str(BASE_PATH / 'Описание отчетов' / 'Шаблон-заявки-на-перемещение-товаров.xlsx'),
        help='Путь к шаблону заявки с маршрутами'
    )
    return parser.parse_args()


# ============================================
# УТИЛИТЫ
# ============================================

def normalize_col(s):
    """Нормализация строки: убираем \\xa0, лишние пробелы"""
    if not isinstance(s, str):
        return s
    return s.replace('\xa0', ' ').strip()


def normalize_wh(s):
    """Нормализация имени склада для сопоставления с маршрутами"""
    if not isinstance(s, str):
        return ''
    return s.strip().upper().replace(' ', '_')


# ============================================
# АВТООБНАРУЖЕНИЕ ФАЙЛОВ
# ============================================

def detect_files(data_path):
    """Находит файлы данных в папке"""
    delivery_file = None
    stock_file = None

    for f in data_path.iterdir():
        name = f.name.lower()
        if f.suffix == '.xlsx':
            if 'average_delivery' in name or 'средн' in name:
                delivery_file = f.name
            elif 'остатк' in name or 'управлени' in name:
                stock_file = f.name

    if not delivery_file:
        raise FileNotFoundError(f"Не найден файл времени доставки в {data_path}")
    if not stock_file:
        raise FileNotFoundError(f"Не найден файл остатков в {data_path}")

    print(f"   Время доставки: {delivery_file}")
    print(f"   Остатки: {stock_file}")
    return delivery_file, stock_file


def extract_period(data_path, delivery_file):
    """Извлекает период из файла доставки"""
    import openpyxl
    wb = openpyxl.load_workbook(data_path / delivery_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row0 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    period_str = str(row0[0]) if row0[0] else ''
    wb.close()

    match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s*[-–]\s*(\d{2}\.\d{2}\.\d{4})', period_str)
    if match:
        return match.group(1), match.group(2)
    return None, None


# ============================================
# ЗАГРУЗКА ДАННЫХ
# ============================================

def load_valid_routes(template_path):
    """Загрузка валидных маршрутов перемещений из справочника"""
    print("0. Загрузка маршрутов из справочника...")
    import openpyxl

    p = Path(template_path)
    if not p.exists():
        print(f"   ВНИМАНИЕ: Шаблон не найден: {p}")
        print("   Маршруты не будут валидироваться")
        return set()

    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb['Справочник ФФ']

    routes = set()
    for row in ws.iter_rows(min_row=40, values_only=True):
        src = row[2]
        dst = row[3]
        if src and dst:
            routes.add((normalize_wh(str(src)), normalize_wh(str(dst))))

    wb.close()
    src_count = len(set(r[0] for r in routes))
    dst_count = len(set(r[1] for r in routes))
    print(f"   Маршрутов: {len(routes)} (складов-отправителей: {src_count}, получателей: {dst_count})")
    return routes


def load_statuses(skip=False):
    """Загрузка статусов артикулов из Supabase"""
    if skip:
        print("   Статусы: пропущено (--no-statuses)")
        return {}

    if not HAS_DATA_LAYER:
        print("   Статусы: data_layer не найден, продолжаем без статусов")
        return {}

    print("   Загрузка статусов из Supabase...")
    statuses = get_artikuly_statuses()

    if not statuses:
        print("   Статусы не загружены")
        return {}

    status_counts = {}
    for s in statuses.values():
        status_counts[s] = status_counts.get(s, 0) + 1

    vyvodim = status_counts.get('Выводим', 0)
    print(f"   Загружено: {len(statuses)} артикулов, 'Выводим': {vyvodim}")
    return statuses


def load_delivery_data(data_path, delivery_file):
    """Загрузка данных времени доставки"""
    print("1. Загрузка данных времени доставки...")
    filepath = data_path / delivery_file

    # Лист "По кластерам": 9 колонок после skiprows=5
    df_clusters = pd.read_excel(filepath, sheet_name='По кластерам',
                                header=None, skiprows=5)
    df_clusters.columns = [
        'Артикул', 'SKU', 'col2', 'Схема',
        'Кластер_отгрузки', 'Кластер_доставки',
        'Нормативное_время', 'Скорость', 'Отгружено'
    ]
    df_clusters = df_clusters.dropna(subset=['Артикул']).copy()
    df_clusters['Отгружено'] = pd.to_numeric(df_clusters['Отгружено'], errors='coerce').fillna(0).astype(int)
    df_clusters['Нормативное_время'] = pd.to_numeric(df_clusters['Нормативное_время'], errors='coerce').fillna(0).astype(int)

    # Лист "По товарам"
    df_products = pd.read_excel(filepath, sheet_name='По товарам',
                                header=None, skiprows=5)
    df_products.columns = [
        'Артикул', 'SKU', 'Название', 'Рекомендуемое_кол',
        'Кластер_доставки', 'Среднее_время', 'Скорость', 'Схема',
        'Отгружено_всего', 'Отгружено_быстро', 'Отгружено_средне',
        'Отгружено_долго', 'Доля_влияния', 'Экономия_руб'
    ]
    df_products = df_products.dropna(subset=['Артикул']).copy()

    def parse_pct(val):
        if pd.isna(val):
            return 0.0
        s = str(val).replace('%', '').replace(',', '.').strip()
        try:
            return float(s)
        except ValueError:
            return 0.0

    df_products['Доля_влияния'] = df_products['Доля_влияния'].apply(parse_pct)
    df_products['Экономия_руб'] = pd.to_numeric(df_products['Экономия_руб'], errors='coerce').fillna(0)
    df_products['Отгружено_всего'] = pd.to_numeric(df_products['Отгружено_всего'], errors='coerce').fillna(0).astype(int)

    print(f"   По кластерам: {len(df_clusters)} маршрутов")
    print(f"   По товарам: {len(df_products)} строк")
    return df_clusters, df_products


def load_normative_matrix(data_path, delivery_file):
    """Загрузка матрицы нормативного времени доставки 25×25"""
    print("2. Загрузка матрицы нормативного времени...")
    import openpyxl
    wb = openpyxl.load_workbook(data_path / delivery_file, data_only=True)
    ws = wb['Нормативный срок доставки']

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    shipping_clusters = [normalize_col(str(v)) for v in header_row[1:] if v is not None]

    matrix = {}
    delivery_clusters = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        dc = normalize_col(str(row[0])) if row[0] else None
        if not dc:
            continue
        delivery_clusters.append(dc)
        values = [v for v in row[1:] if v is not None]
        for j, sc in enumerate(shipping_clusters):
            if j < len(values):
                try:
                    matrix[(dc, sc)] = int(values[j])
                except (ValueError, TypeError):
                    pass

    wb.close()
    print(f"   Кластеров: {len(delivery_clusters)} × {len(shipping_clusters)}")
    return matrix, delivery_clusters, shipping_clusters


def load_stock_data(data_path, stock_file):
    """Загрузка остатков: Товар-кластер + Товар-склад + Кластеры"""
    print("3. Загрузка данных остатков...")
    filepath = data_path / stock_file

    # Лист "Товар-кластер"
    df_tc = pd.read_excel(filepath, sheet_name='Товар-кластер',
                          header=None, skiprows=4)
    df_tc.columns = [
        'Артикул', 'Название', 'SKU', 'Признак', 'Зона',
        'Кластер', 'Статус', 'Дней_до_конца', 'Среднесуточные_продажи',
        'Дней_без_продаж', 'Доступно', 'Готовим', 'Маркир_вывоз',
        'Маркир_УПД', 'Срок_годности', 'Брак_поставка', 'Брак_сток',
        'Излишки', 'Проверка', 'Заявки_поставка', 'В_пути', 'Возвраты', 'Вывоз'
    ]
    df_tc = df_tc.dropna(subset=['Артикул']).copy()
    df_tc['Кластер'] = df_tc['Кластер'].apply(normalize_col)
    df_tc['Статус'] = df_tc['Статус'].apply(normalize_col)
    df_tc['Доступно'] = pd.to_numeric(df_tc['Доступно'], errors='coerce').fillna(0).astype(int)
    df_tc['Среднесуточные_продажи'] = pd.to_numeric(df_tc['Среднесуточные_продажи'], errors='coerce').fillna(0)

    def parse_days(val):
        if pd.isna(val):
            return 0
        s = str(val).lower()
        if 'больше' in s:
            digits = re.findall(r'\d+', s)
            return int(digits[0]) + 1 if digits else 121
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    df_tc['Дней_без_продаж'] = df_tc['Дней_без_продаж'].apply(parse_days)

    # Лист "Товар-склад" — для маппинга артикул→склад→кластер
    df_tw = pd.read_excel(filepath, sheet_name='Товар-склад',
                          header=None, skiprows=4)
    df_tw.columns = [
        'Артикул', 'Название', 'SKU', 'Признаки', 'Зона', 'Кластер', 'Склад',
        'Доступно', 'Скоро', 'Снято', 'Проблемы', 'Истек', 'Дефекты', 'Поврежд',
        'Неопозн', 'Проверка', 'Заявки', 'Доставка', 'Возвраты', 'Вывоз'
    ]
    df_tw = df_tw.dropna(subset=['Артикул']).copy()
    df_tw['Доступно'] = pd.to_numeric(df_tw['Доступно'], errors='coerce').fillna(0).astype(int)
    df_tw['Склад_upper'] = df_tw['Склад'].apply(normalize_wh)
    df_tw['Кластер'] = df_tw['Кластер'].apply(normalize_col)

    # Лист "Кластеры"
    df_cl = pd.read_excel(filepath, sheet_name='Кластеры',
                          header=None, skiprows=4)
    df_cl.columns = [
        'Кластер', 'Статус', 'Дней_до_конца', 'Среднесуточные_продажи',
        'Дней_без_продаж', 'Доступно', 'Готовим', 'Маркир_вывоз',
        'Маркир_УПД', 'Срок_годности', 'Брак_поставка', 'Брак_сток',
        'Излишки', 'Проверка', 'Заявки_поставка', 'В_пути', 'Возвраты', 'Вывоз'
    ]
    df_cl = df_cl.dropna(subset=['Кластер']).copy()
    df_cl['Кластер'] = df_cl['Кластер'].apply(normalize_col)
    df_cl['Статус'] = df_cl['Статус'].apply(normalize_col)
    df_cl['Доступно'] = pd.to_numeric(df_cl['Доступно'], errors='coerce').fillna(0).astype(int)
    df_cl['Среднесуточные_продажи'] = pd.to_numeric(df_cl['Среднесуточные_продажи'], errors='coerce').fillna(0)

    print(f"   Товар-кластер: {len(df_tc)} строк")
    print(f"   Товар-склад: {len(df_tw)} строк, всего {df_tw['Доступно'].sum()} шт")
    print(f"   Кластеры: {len(df_cl)} строк")

    return df_tc, df_tw, df_cl


# ============================================
# РАСЧЁТ МЕТРИК
# ============================================

def calculate_article_stats(df_clusters, target_time):
    """Расчёт средневзвешенного времени доставки по артикулам"""
    print("4. Расчёт метрик по артикулам...")

    df = df_clusters[
        (df_clusters['Схема'] == 'FBO') &
        (df_clusters['Отгружено'] > 0)
    ].copy()

    stats = df.groupby('Артикул').agg(
        SKU=('SKU', 'first'),
        Название=('col2', 'first'),  # col2 содержит числовое значение, Название в products
        Всего_отгружено=('Отгружено', 'sum'),
        Время_взвешенное=('Нормативное_время',
                          lambda x: (x * df.loc[x.index, 'Отгружено']).sum()),
    ).reset_index()

    stats['Среднее_время'] = np.where(
        stats['Всего_отгружено'] > 0,
        np.ceil(stats['Время_взвешенное'] / stats['Всего_отгружено']).astype(int),
        0
    )

    speed_counts = df.groupby('Артикул').apply(
        lambda g: pd.Series({
            'Быстро_шт': g.loc[g['Нормативное_время'] <= 28, 'Отгружено'].sum(),
            'Средне_шт': g.loc[
                (g['Нормативное_время'] > 28) & (g['Нормативное_время'] <= 75),
                'Отгружено'].sum(),
            'Долго_шт': g.loc[g['Нормативное_время'] > 75, 'Отгружено'].sum(),
        })
    ).reset_index()

    stats = stats.merge(speed_counts, on='Артикул', how='left')
    stats[['Быстро_шт', 'Средне_шт', 'Долго_шт']] = (
        stats[['Быстро_шт', 'Средне_шт', 'Долго_шт']].fillna(0).astype(int)
    )

    stats['Доля_быстро_%'] = np.where(
        stats['Всего_отгружено'] > 0,
        (stats['Быстро_шт'] / stats['Всего_отгружено'] * 100).round(1),
        0
    )

    def get_status(t):
        if t <= target_time:
            return 'OK'
        elif t <= ATTENTION_TIME:
            return 'Внимание'
        return 'Критично'

    stats['Статус'] = stats['Среднее_время'].apply(get_status)
    stats = stats.drop(columns=['Время_взвешенное'])
    stats = stats.sort_values('Среднее_время', ascending=False)

    ok = len(stats[stats['Статус'] == 'OK'])
    att = len(stats[stats['Статус'] == 'Внимание'])
    crit = len(stats[stats['Статус'] == 'Критично'])
    print(f"   Артикулов: {len(stats)} | OK: {ok} | Внимание: {att} | Критично: {crit}")

    total_qty = stats['Всего_отгружено'].sum()
    overall_time = (
        math.ceil((stats['Среднее_время'] * stats['Всего_отгружено']).sum() / total_qty)
        if total_qty > 0 else 0
    )
    print(f"   Общее средневзвешенное время: {overall_time} ч (цель: ≤{target_time} ч)")

    return stats, overall_time


# ============================================
# ЗАЩИТА ДОНОРОВ
# ============================================

def calculate_donor_limit(stock, daily_sales, status, safety_days):
    """Рассчитывает доступное для трансфера количество"""
    if stock <= 0:
        return 0, 0, 'нет остатков'

    norm_status = normalize_col(status)

    if norm_status in BLOCKED_STATUSES:
        return 0, stock, f'статус: {norm_status}'

    if daily_sales == 0:
        if 'без продаж' in norm_status.lower() or 'избыточн' in norm_status.lower():
            return stock, 0, 'нет спроса'

    if norm_status in CAUTIOUS_STATUSES:
        safe_min = math.ceil(daily_sales * safety_days * 1.5)
    else:
        safe_min = math.ceil(daily_sales * safety_days)

    available = max(0, stock - safe_min)
    factor = 'буфер безопасности' if available == 0 else 'доступен'
    return available, safe_min, factor


# ============================================
# ГЕНЕРАЦИЯ ПЕРЕМЕЩЕНИЙ (v2 — склад→склад)
# ============================================

def generate_movements(article_stats, df_clusters, df_stock_tc, df_stock_tw,
                       normative_matrix, all_clusters,
                       valid_routes, vyvodim_articles,
                       target_time, safety_days, avg_price,
                       movement_cost_per_unit, planning_weeks):
    """
    Генерация перемещений на уровне СКЛАДОВ с валидацией маршрутов.

    Только реальный спрос: перемещаем артикулы, у которых есть спрос
    на целевом кластере по данным доставки (>29ч маршруты).

    Маршруты валидируются по справочнику OZON.
    Артикулы в статусе «Выводим» исключаются.
    """
    print("5. Генерация перемещений (склад→склад, реальный спрос)...")

    movements = []
    supplies = []
    donor_records = []

    # Глобальный трекер доступности доноров (кластер-уровень)
    donor_capacity = {}
    donor_records_seen = set()

    def get_donor_capacity(article, cluster):
        key = (article, cluster)
        if key not in donor_capacity:
            row = df_stock_tc[
                (df_stock_tc['Артикул'] == article) &
                (df_stock_tc['Кластер'] == cluster)
            ]
            if len(row) == 0:
                donor_capacity[key] = (0, 0, 'нет данных')
                return 0, 0, 'нет данных'
            r = row.iloc[0]
            avail, safe_min, factor = calculate_donor_limit(
                int(r['Доступно']), float(r['Среднесуточные_продажи']),
                str(r['Статус']), safety_days
            )
            donor_capacity[key] = (avail, safe_min, factor)

            if key not in donor_records_seen:
                donor_records_seen.add(key)
                donor_records.append({
                    'Артикул': article,
                    'Кластер_донор': cluster,
                    'Текущие_остатки': int(r['Доступно']),
                    'Среднесуточные_продажи': float(r['Среднесуточные_продажи']),
                    'Статус': str(r['Статус']),
                    'Безопасный_минимум': safe_min,
                    'Доступно_для_трансфера': avail,
                    'Фактор_ограничения': factor,
                })

        return donor_capacity[key]

    consumed = {}

    def consume_donor(article, cluster, qty):
        avail, _, _ = get_donor_capacity(article, cluster)
        key = (article, cluster)
        already = consumed.get(key, 0)
        actual = min(qty, avail - already)
        if actual > 0:
            consumed[key] = already + actual
        return max(0, actual)

    # Маппинг склад → кластер из stock data
    wh_to_cluster = {}
    for _, r in df_stock_tw.iterrows():
        wh = r['Склад_upper']
        cl = r['Кластер']
        if wh and cl:
            wh_to_cluster[wh] = cl

    # Маппинг кластер → склады, которые могут ОТПРАВЛЯТЬ
    cluster_to_src_warehouses = {}
    for wh, cl in wh_to_cluster.items():
        if any(wh == r[0] for r in valid_routes):
            cluster_to_src_warehouses.setdefault(cl, set()).add(wh)

    # Артикулы со спросом на каждом кластере (медленные маршруты)
    # Спрос = маршруты, где Нормативное_время > target и есть отгрузки
    df_fbo = df_clusters[
        (df_clusters['Схема'] == 'FBO') &
        (df_clusters['Отгружено'] > 0)
    ].copy()

    df_slow = df_fbo[df_fbo['Нормативное_время'] > target_time].copy()

    # Проблемные артикулы
    problem = article_stats[article_stats['Среднее_время'] > target_time].copy()
    problem['impact'] = problem['Всего_отгружено'] * (problem['Среднее_время'] - target_time)
    problem = problem.sort_values('impact', ascending=False)

    # Фильтр «Выводим»
    vyvodim_excluded = 0

    for _, row in problem.iterrows():
        article = row['Артикул']

        if article.lower() in vyvodim_articles:
            vyvodim_excluded += 1
            continue

        avg_time = int(row['Среднее_время'])
        total_shipped = int(row['Всего_отгружено'])

        # Медленные маршруты этого артикула
        slow_routes = df_slow[df_slow['Артикул'] == article].copy()
        if len(slow_routes) == 0:
            continue

        # Группировка по кластеру доставки (куда идёт спрос)
        delivery_demand = slow_routes.groupby('Кластер_доставки').agg(
            Спрос_нед=('Отгружено', 'sum'),
            Средн_время=('Нормативное_время',
                         lambda x: math.ceil(
                             (x * slow_routes.loc[x.index, 'Отгружено']).sum()
                             / slow_routes.loc[x.index, 'Отгружено'].sum()
                         ) if slow_routes.loc[x.index, 'Отгружено'].sum() > 0 else 0),
        ).reset_index()

        # Целевые кластеры
        target_clusters = set(
            normalize_col(str(c)) for c in delivery_demand['Кластер_доставки']
        )

        # Кол-во = спрос × planning_weeks (реальная потребность)
        delivery_demand['Нужно'] = (delivery_demand['Спрос_нед'] * planning_weeks).astype(int)
        delivery_demand = delivery_demand.sort_values('Нужно', ascending=False)

        for _, dd in delivery_demand.iterrows():
            delivery_cluster = normalize_col(str(dd['Кластер_доставки']))
            needed_qty = int(dd['Нужно'])
            weekly_demand = int(dd['Спрос_нед'])
            current_time = int(dd['Средн_время'])

            if needed_qty <= 0 or delivery_cluster in INTERNATIONAL_CLUSTERS:
                continue

            remaining = needed_qty

            # Ищем доноров: кластеры с излишком этого артикула
            for cluster in all_clusters:
                nc = normalize_col(str(cluster))
                if nc in target_clusters or nc in INTERNATIONAL_CLUSTERS:
                    continue

                avail_capacity, _, _ = get_donor_capacity(article, nc)
                already_used = consumed.get((article, nc), 0)
                available = avail_capacity - already_used

                if available <= 0:
                    continue

                transfer_qty = min(remaining, available)

                # Найти конкретный склад-отправитель и склад-получатель
                src_warehouses = cluster_to_src_warehouses.get(nc, set())
                dst_wh = None
                for d_wh, d_cl in DEST_WAREHOUSE_TO_CLUSTER.items():
                    if d_cl == delivery_cluster:
                        dst_wh = d_wh
                        break

                if not dst_wh:
                    continue

                # Проверяем есть ли валидный маршрут
                matched_src_wh = None
                if valid_routes:
                    for swh in src_warehouses:
                        if (swh, dst_wh) in valid_routes:
                            matched_src_wh = swh
                            break
                else:
                    # Нет справочника — берём первый
                    matched_src_wh = next(iter(src_warehouses)) if src_warehouses else None

                if not matched_src_wh:
                    continue

                actual = consume_donor(article, nc, transfer_qty)
                if actual <= 0:
                    continue

                surcharge = SURCHARGE_RATES.get(delivery_cluster, 0)
                weekly_savings = round(avg_price * surcharge * weekly_demand, 2) if surcharge > 0 else 0

                movements.append({
                    'Артикул': article,
                    'SKU': row.get('SKU', ''),
                    'Название': row.get('Название', ''),
                    'Откуда_кластер': nc,
                    'Откуда_склад': matched_src_wh,
                    'Куда_кластер': delivery_cluster,
                    'Куда_склад': dst_wh,
                    'Кол_во': actual,
                    'Спрос_нед': weekly_demand,
                    'Время_текущее': current_time,
                    'Время_новое': 28,
                    'Экономия_времени': current_time - 28,
                    'Среднее_время_артикула': avg_time,
                    'Всего_отгружено': total_shipped,
                    'Экономия_наценка_нед_руб': weekly_savings,
                    'Стоимость_перемещения_руб': round(actual * movement_cost_per_unit, 2),
                })

                remaining -= actual
                if remaining <= 0:
                    break

            if remaining > 0 and remaining >= needed_qty * 0.5:
                supplies.append({
                    'Артикул': article,
                    'SKU': row.get('SKU', ''),
                    'Название': row.get('Название', ''),
                    'Куда_кластер': delivery_cluster,
                    'Кол_во': remaining,
                    'Спрос_нед': weekly_demand,
                    'Текущее_время': current_time,
                    'Время_при_локальном': 28,
                    'Экономия_времени': current_time - 28,
                    'Причина': 'Недостаточно доноров / нет маршрута',
                })

    moves_df = pd.DataFrame(movements) if movements else pd.DataFrame()
    supply_df = pd.DataFrame(supplies) if supplies else pd.DataFrame()
    donor_df = pd.DataFrame(donor_records) if donor_records else pd.DataFrame()

    total_units = moves_df['Кол_во'].sum() if not moves_df.empty else 0
    supply_units = supply_df['Кол_во'].sum() if not supply_df.empty else 0
    print(f"   Перемещений: {len(moves_df)} ({total_units} шт)")
    print(f"   Допоставок: {len(supply_df)} ({supply_units} шт)")
    if vyvodim_excluded > 0:
        print(f"   Исключено 'Выводим': {vyvodim_excluded} артикулов")
    if len(donor_df) > 0:
        limited = donor_df[donor_df['Доступно_для_трансфера'] == 0]
        print(f"   Доноров с ограничением: {len(limited)} из {len(donor_df)}")

    return moves_df, supply_df, donor_df


# ============================================
# АНАЛИЗ СКЛАДОВ (НОВЫЙ)
# ============================================

def analyze_warehouse_transfers(moves_df, df_stock_tw, valid_routes, vyvodim_articles):
    """
    Анализ возможных перемещений на уровне складов.
    Для каждого склада-отправителя: что можно отправить и куда.
    """
    print("   Анализ складов...")

    if moves_df.empty:
        return pd.DataFrame()

    # Агрегация по парам склад→склад
    wh_agg = moves_df.groupby(['Откуда_склад', 'Куда_склад', 'Откуда_кластер', 'Куда_кластер']).agg(
        Артикулов=('Артикул', 'nunique'),
        Всего_единиц=('Кол_во', 'sum'),
        Спрос_нед_итого=('Спрос_нед', 'sum'),
        Экономия_наценка=('Экономия_наценка_нед_руб', 'sum'),
        Стоимость=('Стоимость_перемещения_руб', 'sum'),
        Средн_экономия_часов=('Экономия_времени', 'mean'),
        Артикулы_список=('Артикул', lambda x: ', '.join(sorted(set(str(v) for v in x)))),
    ).reset_index()

    wh_agg['Всего_единиц'] = wh_agg['Всего_единиц'].astype(int)
    wh_agg['Средн_экономия_часов'] = wh_agg['Средн_экономия_часов'].round(0).astype(int)

    wh_agg['Статус'] = wh_agg['Всего_единиц'].apply(
        lambda q: 'Можно подать заявку' if q >= OZON_MIN_MOVEMENT
        else f'Не хватает {OZON_MIN_MOVEMENT - q} шт'
    )

    wh_agg = wh_agg.sort_values('Всего_единиц', ascending=False)
    return wh_agg


def analyze_fbo_supply_plan(df_clusters, df_stock_tc, article_stats, vyvodim_articles,
                            target_time, valid_routes):
    """
    План FBO-поставки: какие артикулы отгрузить на какие кластеры.
    Рекомендации для новых поставок (не перемещений).
    """
    print("   Формирование плана FBO-поставок...")

    # Артикулы с проблемным временем
    problem = article_stats[article_stats['Среднее_время'] > target_time].copy()

    df_fbo = df_clusters[
        (df_clusters['Схема'] == 'FBO') &
        (df_clusters['Отгружено'] > 0) &
        (df_clusters['Нормативное_время'] > target_time)
    ].copy()

    # Спрос по кластерам доставки
    demand = df_fbo.groupby('Кластер_доставки')['Отгружено'].sum().sort_values(ascending=False)

    # Для каждого целевого кластера: какие артикулы нужны
    rows = []
    for cluster, total_demand in demand.items():
        nc = normalize_col(str(cluster))
        if nc in INTERNATIONAL_CLUSTERS:
            continue

        # Артикулы со спросом на этом кластере
        arts = df_fbo[df_fbo['Кластер_доставки'] == cluster].groupby('Артикул').agg(
            Спрос_нед=('Отгружено', 'sum'),
            Средн_время=('Нормативное_время', 'mean'),
        ).reset_index()

        # Текущий остаток на этом кластере
        stock_on_cluster = df_stock_tc[df_stock_tc['Кластер'] == nc].set_index('Артикул')['Доступно'].to_dict()

        for _, a in arts.iterrows():
            article = a['Артикул']
            if article.lower() in vyvodim_articles:
                continue

            current_stock = stock_on_cluster.get(article, 0)
            weekly = int(a['Спрос_нед'])

            rows.append({
                'Куда_кластер': nc,
                'Артикул': article,
                'Спрос_нед': weekly,
                'Текущий_остаток': current_stock,
                'Средн_время': int(a['Средн_время']),
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['Куда_кластер', 'Спрос_нед'], ascending=[True, False])

    # Сводка по кластерам
    if not df.empty:
        cluster_plan = df.groupby('Куда_кластер').agg(
            Артикулов=('Артикул', 'nunique'),
            Суммарный_спрос_нед=('Спрос_нед', 'sum'),
        ).reset_index().sort_values('Суммарный_спрос_нед', ascending=False)
    else:
        cluster_plan = pd.DataFrame()

    print(f"   Кластеров для поставки: {len(cluster_plan)}")
    return df, cluster_plan


# ============================================
# ПРИОРИТИЗАЦИЯ
# ============================================

def add_priority(df, time_col='Среднее_время_артикула', qty_col='Всего_отгружено'):
    """Добавление приоритета"""
    if df.empty or time_col not in df.columns:
        return df

    def get_priority(row):
        t = row.get(time_col, 0)
        qty = row.get(qty_col, 0)
        if pd.isna(t):
            return 'Нет данных', 0
        if t > CRITICAL_TIME and qty >= HIGH_VOLUME:
            return 'Критично+', 35
        elif t > CRITICAL_TIME:
            return 'Критично', 30
        elif t > ATTENTION_TIME and qty >= HIGH_VOLUME:
            return 'Важно+', 25
        elif t > TARGET_DELIVERY_TIME and qty >= HIGH_VOLUME:
            return 'Важно', 20
        elif t > TARGET_DELIVERY_TIME:
            return 'Можно улучшить', 15
        return 'OK', 10

    df = df.copy()
    df['Приоритет'], df['Балл'] = zip(*df.apply(get_priority, axis=1))
    return df.sort_values('Балл', ascending=False)


def split_by_priority(df):
    if df.empty or 'Приоритет' not in df.columns:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    critical = df[df['Приоритет'].str.contains('Критично', na=False)]
    important = df[df['Приоритет'].str.contains('Важно', na=False)]
    other = df[~df.index.isin(critical.index) & ~df.index.isin(important.index)]
    return critical, important, other


# ============================================
# АГРЕГАЦИЯ
# ============================================

def aggregate_by_direction(moves_df):
    """Агрегация по направлениям склад→склад"""
    if moves_df.empty:
        return pd.DataFrame()

    agg = moves_df.groupby(['Откуда_склад', 'Куда_склад', 'Откуда_кластер', 'Куда_кластер']).agg(
        Артикулов=('Артикул', 'nunique'),
        Всего_единиц=('Кол_во', 'sum'),
        Экономия_наценка=('Экономия_наценка_нед_руб', 'sum'),
        Стоимость=('Стоимость_перемещения_руб', 'sum'),
        Средн_экономия_ч=('Экономия_времени', 'mean'),
        Артикулы=('Артикул', lambda x: ', '.join(sorted(set(str(v) for v in x)))),
    ).reset_index()

    agg['Всего_единиц'] = agg['Всего_единиц'].astype(int)
    agg['Средн_экономия_ч'] = agg['Средн_экономия_ч'].round(0).astype(int)
    agg['Не_хватает_до_200'] = np.maximum(0, OZON_MIN_MOVEMENT - agg['Всего_единиц'])
    agg['Статус_заявки'] = agg['Всего_единиц'].apply(
        lambda q: 'Можно подать заявку' if q >= OZON_MIN_MOVEMENT
        else f'Не хватает {OZON_MIN_MOVEMENT - q} шт'
    )

    return agg.sort_values('Всего_единиц', ascending=False)


def aggregate_supplies_by_cluster(supply_df):
    if supply_df.empty:
        return pd.DataFrame()

    agg = supply_df.groupby('Куда_кластер').agg(
        Артикулов=('Артикул', 'nunique'),
        Всего_единиц=('Кол_во', 'sum'),
        Средн_экономия_ч=('Экономия_времени', 'mean'),
        Артикулы=('Артикул', lambda x: ', '.join(sorted(set(str(v) for v in x)))),
    ).reset_index()
    agg['Всего_единиц'] = agg['Всего_единиц'].astype(int)
    agg['Средн_экономия_ч'] = agg['Средн_экономия_ч'].round(0).astype(int)
    return agg.sort_values('Всего_единиц', ascending=False)


# ============================================
# СВОДКА ПО КЛАСТЕРАМ
# ============================================

def calc_cluster_summary(df_clusters, df_cl, normative_matrix, all_clusters):
    print("6. Сводка по кластерам...")

    demand = df_clusters[
        (df_clusters['Схема'] == 'FBO') &
        (df_clusters['Отгружено'] > 0)
    ].groupby('Кластер_доставки').agg(
        Спрос_шт=('Отгружено', 'sum'),
        Средн_время=('Нормативное_время', lambda x: (
            math.ceil((x * df_clusters.loc[x.index, 'Отгружено']).sum()
                       / df_clusters.loc[x.index, 'Отгружено'].sum())
            if df_clusters.loc[x.index, 'Отгружено'].sum() > 0 else 0
        )),
    ).reset_index()
    demand.columns = ['Кластер', 'Спрос_шт', 'Средн_время_в_кластер']

    summary = df_cl[['Кластер', 'Статус', 'Среднесуточные_продажи', 'Доступно', 'Дней_до_конца']].copy()
    summary = summary.merge(demand, on='Кластер', how='left')
    summary['Спрос_шт'] = summary['Спрос_шт'].fillna(0).astype(int)
    summary['Средн_время_в_кластер'] = summary['Средн_время_в_кластер'].fillna(0).astype(int)
    summary['Наценка_%'] = summary['Кластер'].map(lambda c: SURCHARGE_RATES.get(c, 0) * 100)
    summary['Международный'] = summary['Кластер'].apply(
        lambda c: 'Да' if c in INTERNATIONAL_CLUSTERS else 'Нет'
    )
    summary = summary.sort_values('Спрос_шт', ascending=False)
    print(f"   Кластеров: {len(summary)}")
    return summary


# ============================================
# ЭКОНОМИЧЕСКИЙ АНАЛИЗ
# ============================================

def calculate_economics(moves_df, df_products):
    if moves_df.empty:
        return moves_df

    df = moves_df.copy()

    savings_map = {}
    if not df_products.empty:
        for _, row in df_products.iterrows():
            key = (str(row['Артикул']), normalize_col(str(row.get('Кластер_доставки', ''))))
            savings_map[key] = row.get('Экономия_руб', 0)

    df['Экономия_OZON_руб_нед'] = df.apply(
        lambda r: savings_map.get((str(r['Артикул']), str(r['Куда_кластер'])), 0), axis=1
    )
    df['Экономия_итого_нед'] = df['Экономия_наценка_нед_руб'] + df['Экономия_OZON_руб_нед']
    df['ROI_годовой'] = np.where(
        df['Стоимость_перемещения_руб'] > 0,
        ((df['Экономия_итого_нед'] * 52) / df['Стоимость_перемещения_руб']).round(2),
        np.inf
    )
    df['Окупаемость_нед'] = np.where(
        df['Экономия_итого_нед'] > 0,
        (df['Стоимость_перемещения_руб'] / df['Экономия_итого_нед']).round(1),
        np.inf
    )
    return df


# ============================================
# СОХРАНЕНИЕ ОТЧЁТА
# ============================================

def save_report(article_stats, moves_df, supply_df, donor_df,
                cluster_summary, direction_agg, supply_cluster_agg,
                fbo_plan, fbo_cluster_plan,
                normative_matrix, delivery_clusters, shipping_clusters,
                overall_time, output_file, args, period_start, period_end,
                statuses, vyvodim_articles):
    """Сохранение отчёта v2"""
    print("7. Сохранение отчёта...")

    article_stats = add_priority(article_stats, 'Среднее_время', 'Всего_отгружено')
    moves_df = add_priority(moves_df) if not moves_df.empty else moves_df
    critical, important, other = split_by_priority(moves_df)

    total_articles = len(article_stats)
    ok_count = len(article_stats[article_stats['Статус'] == 'OK'])
    att_count = len(article_stats[article_stats['Статус'] == 'Внимание'])
    crit_count = len(article_stats[article_stats['Статус'] == 'Критично'])

    total_move_units = moves_df['Кол_во'].sum() if not moves_df.empty else 0
    total_supply_units = supply_df['Кол_во'].sum() if not supply_df.empty else 0
    total_savings = (moves_df['Экономия_итого_нед'].sum()
                     if not moves_df.empty and 'Экономия_итого_нед' in moves_df.columns else 0)
    total_move_cost = moves_df['Стоимость_перемещения_руб'].sum() if not moves_df.empty else 0

    # Прогнозное время
    if not moves_df.empty:
        total_qty = article_stats['Всего_отгружено'].sum()
        time_saved = (moves_df['Экономия_времени'] * moves_df['Спрос_нед']).sum()
        projected_time = (
            max(28, math.ceil((overall_time * total_qty - time_saved) / total_qty))
            if total_qty > 0 else overall_time
        )
    else:
        projected_time = overall_time

    donors_limited = 0
    if not donor_df.empty and 'Доступно_для_трансфера' in donor_df.columns:
        donors_limited = len(donor_df[donor_df['Доступно_для_трансфера'] == 0])

    dir_ready = 0
    dir_total = 0
    if not direction_agg.empty:
        dir_total = len(direction_agg)
        dir_ready = len(direction_agg[direction_agg['Всего_единиц'] >= OZON_MIN_MOVEMENT])

    period_text = f'{period_start} – {period_end}' if period_start else 'Не определён'

    summary_rows = [
        ['=== ОБЩАЯ СТАТИСТИКА ===', '', ''],
        ['Период данных', period_text, ''],
        ['Всего артикулов FBO', total_articles, ''],
        ['Текущее средн. время доставки', f'{overall_time} ч', f'Цель: ≤{args.target_time} ч'],
        ['Прогнозное время после оптимизации', f'{projected_time} ч', ''],
        ['Горизонт планирования', f'{args.planning_weeks} нед.', ''],
        ['', '', ''],
        ['=== РАСПРЕДЕЛЕНИЕ ===', '', ''],
        ['OK (≤29 ч)', f'{ok_count} ({ok_count*100//max(total_articles,1)}%)', ''],
        ['Внимание (30-60 ч)', f'{att_count} ({att_count*100//max(total_articles,1)}%)', ''],
        ['Критично (>60 ч)', f'{crit_count} ({crit_count*100//max(total_articles,1)}%)', ''],
        ['', '', ''],
        ['=== ПЕРЕМЕЩЕНИЯ (только реальный спрос) ===', '', ''],
        ['Критичных', f'{len(critical)} ({critical["Кол_во"].sum() if not critical.empty else 0} шт)', ''],
        ['Важных', f'{len(important)} ({important["Кол_во"].sum() if not important.empty else 0} шт)', ''],
        ['Остальных', f'{len(other)} ({other["Кол_во"].sum() if not other.empty else 0} шт)', ''],
        ['Всего перемещений', f'{len(moves_df)} строк, {total_move_units} шт', ''],
        ['', '', ''],
        ['=== НАПРАВЛЕНИЯ СКЛАД→СКЛАД ===', '', ''],
        ['Всего направлений', dir_total, ''],
        ['Готовы к заявке (≥200 шт)', dir_ready, 'Можно подать заявку в OZON'],
        ['Мин. 200, макс. 10 000 шт на заявку', '', 'Ограничения OZON'],
        ['', '', ''],
        ['=== ДОПОСТАВКИ ===', '', ''],
        ['Допоставок', f'{len(supply_df)} позиций, {total_supply_units} шт', 'Нужна новая FBO-поставка'],
        ['', '', ''],
        ['=== ЭКОНОМИКА ===', '', ''],
        ['Экономия в нед.', f'{total_savings:,.0f} руб', ''],
        ['Экономия в год', f'{total_savings * 52:,.0f} руб', ''],
        ['Стоимость перемещений', f'{total_move_cost:,.0f} руб', ''],
        ['Цена товара (расчёт)', f'{args.avg_product_price} руб', ''],
        ['Стоимость перемещения/ед.', f'{args.movement_cost_per_unit} руб', ''],
        ['', '', ''],
        ['=== ЗАЩИТА ДОНОРОВ ===', '', ''],
        ['Буфер, дней', args.safety_days, ''],
        ['Доноров с ограничением', donors_limited, ''],
    ]

    if statuses:
        vyvodim_total = sum(1 for s in statuses.values() if s == 'Выводим')
        summary_rows.extend([
            ['', '', ''],
            ['=== СТАТУСЫ (Wookiee SKU Database) ===', '', ''],
            ['Всего артикулов в базе', len(statuses), ''],
            ['Статус "Выводим"', vyvodim_total, 'Исключены из перемещений'],
        ])

    summary = pd.DataFrame(summary_rows, columns=['Показатель', 'Значение', 'Примечание'])

    move_cols = [
        'Артикул', 'SKU', 'Название',
        'Откуда_кластер', 'Откуда_склад', 'Куда_кластер', 'Куда_склад',
        'Кол_во', 'Спрос_нед', 'Время_текущее', 'Время_новое', 'Экономия_времени',
        'Среднее_время_артикула', 'Всего_отгружено',
        'Стоимость_перемещения_руб', 'Приоритет', 'Балл',
    ]

    def fc(df, cols):
        return df[[c for c in cols if c in df.columns]]

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as w:
        summary.to_excel(w, sheet_name='Сводка', index=False)

        if len(critical) > 0:
            fc(critical, move_cols).to_excel(w, sheet_name='1_Критичные', index=False)
        if len(important) > 0:
            fc(important, move_cols).to_excel(w, sheet_name='2_Важные', index=False)
        if len(other) > 0:
            fc(other, move_cols).to_excel(w, sheet_name='3_Остальные', index=False)

        if not moves_df.empty:
            fc(moves_df, move_cols).to_excel(w, sheet_name='Все перемещения', index=False)

        if not direction_agg.empty:
            direction_agg.to_excel(w, sheet_name='Направления', index=False)

        if not supply_df.empty:
            supply_cols = ['Артикул', 'SKU', 'Название', 'Куда_кластер',
                           'Кол_во', 'Спрос_нед', 'Текущее_время',
                           'Время_при_локальном', 'Экономия_времени', 'Причина']
            fc(supply_df, supply_cols).to_excel(w, sheet_name='Допоставки', index=False)

        if not supply_cluster_agg.empty:
            supply_cluster_agg.to_excel(w, sheet_name='Допоставки_кластеры', index=False)

        # План FBO-поставок
        if not fbo_cluster_plan.empty:
            fbo_cluster_plan.to_excel(w, sheet_name='План_FBO_поставок', index=False)
        if not fbo_plan.empty:
            fbo_plan.to_excel(w, sheet_name='Детали_FBO_поставок', index=False)

        # Анализ артикулов
        cols_art = ['Приоритет', 'Артикул', 'SKU', 'Название', 'Всего_отгружено',
                    'Среднее_время', 'Быстро_шт', 'Средне_шт', 'Долго_шт',
                    'Доля_быстро_%', 'Статус', 'Балл']
        article_stats[[c for c in cols_art if c in article_stats.columns]].to_excel(
            w, sheet_name='Анализ_артикулов', index=False)

        cluster_summary.to_excel(w, sheet_name='Анализ_кластеров', index=False)

        if not donor_df.empty:
            donor_df.to_excel(w, sheet_name='Защита доноров', index=False)

        # Матрица
        matrix_data = []
        for dc in delivery_clusters:
            row_data = {'Кластер': dc}
            for sc in shipping_clusters:
                row_data[sc] = normative_matrix.get((dc, sc), '')
            matrix_data.append(row_data)
        if matrix_data:
            pd.DataFrame(matrix_data).to_excel(w, sheet_name='Матрица времени', index=False)

        # Экономика
        if not moves_df.empty and 'Экономия_итого_нед' in moves_df.columns:
            econ_cols = ['Артикул', 'SKU', 'Название',
                         'Откуда_кластер', 'Откуда_склад', 'Куда_кластер', 'Куда_склад',
                         'Кол_во', 'Спрос_нед', 'Время_текущее', 'Время_новое',
                         'Экономия_времени', 'Экономия_наценка_нед_руб',
                         'Экономия_OZON_руб_нед', 'Экономия_итого_нед',
                         'Стоимость_перемещения_руб', 'ROI_годовой', 'Окупаемость_нед']
            fc(moves_df, econ_cols).to_excel(w, sheet_name='Экономика', index=False)

    print(f"   Файл: {output_file}")
    return output_file


# ============================================
# MAIN
# ============================================

def main():
    args = parse_args()

    print("=" * 60)
    print("Отчёт по оптимизации времени доставки OZON (v2)")
    print(f"Данные: {args.data_folder}")
    print(f"Целевое время: {args.target_time} ч")
    print(f"Горизонт: {args.planning_weeks} нед. | Буфер: {args.safety_days} дней")
    print(f"Стоимость перемещения: {args.movement_cost_per_unit} руб/ед.")
    print("=" * 60)

    data_path = BASE_PATH / args.data_folder
    if not data_path.exists():
        print(f"ОШИБКА: Папка не найдена: {data_path}")
        return None

    output_dir = BASE_PATH / 'Отчеты готовые'

    # 0. Маршруты и статусы
    valid_routes = load_valid_routes(args.template)
    statuses = load_statuses(skip=args.no_statuses)
    vyvodim_articles = set(art for art, s in statuses.items() if s == 'Выводим')

    # 1. Файлы
    delivery_file, stock_file = detect_files(data_path)
    period_start, period_end = extract_period(data_path, delivery_file)
    if period_start:
        print(f"   Период: {period_start} – {period_end}")

    # 2. Загрузка данных
    df_clusters, df_products = load_delivery_data(data_path, delivery_file)
    normative_matrix, delivery_clusters, shipping_clusters = load_normative_matrix(data_path, delivery_file)
    df_stock_tc, df_stock_tw, df_cl = load_stock_data(data_path, stock_file)

    df_clusters['Кластер_отгрузки'] = df_clusters['Кластер_отгрузки'].apply(normalize_col)
    df_clusters['Кластер_доставки'] = df_clusters['Кластер_доставки'].apply(normalize_col)
    df_products['Кластер_доставки'] = df_products['Кластер_доставки'].apply(normalize_col)

    all_clusters = list(set(delivery_clusters) | set(shipping_clusters))

    # 3. Метрики
    article_stats, overall_time = calculate_article_stats(df_clusters, args.target_time)

    # 4. Перемещения (только реальный спрос, валидные маршруты)
    moves_df, supply_df, donor_df = generate_movements(
        article_stats, df_clusters, df_stock_tc, df_stock_tw,
        normative_matrix, all_clusters,
        valid_routes, vyvodim_articles,
        args.target_time, args.safety_days,
        args.avg_product_price, args.movement_cost_per_unit,
        args.planning_weeks
    )

    # 5. Экономика
    if not moves_df.empty:
        moves_df = calculate_economics(moves_df, df_products)

    # 6. Сводка по кластерам
    cluster_summary = calc_cluster_summary(df_clusters, df_cl, normative_matrix, all_clusters)

    # 7. Агрегация по направлениям (склад→склад)
    direction_agg = aggregate_by_direction(moves_df)
    supply_cluster_agg = aggregate_supplies_by_cluster(supply_df)

    # 8. План FBO-поставок
    fbo_plan, fbo_cluster_plan = analyze_fbo_supply_plan(
        df_clusters, df_stock_tc, article_stats, vyvodim_articles,
        args.target_time, valid_routes
    )

    # 9. Сохранение
    date_str = datetime.now().strftime('%d-%m-%Y')
    output_file = output_dir / f'Отчет_время_доставки_OZON_{date_str}.xlsx'

    save_report(
        article_stats, moves_df, supply_df, donor_df,
        cluster_summary, direction_agg, supply_cluster_agg,
        fbo_plan, fbo_cluster_plan,
        normative_matrix, delivery_clusters, shipping_clusters,
        overall_time, output_file, args,
        period_start, period_end,
        statuses, vyvodim_articles
    )

    print("=" * 60)
    print("Готово!")
    return output_file


if __name__ == "__main__":
    main()
