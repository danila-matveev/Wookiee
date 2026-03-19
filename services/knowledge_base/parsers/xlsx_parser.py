"""
XLSX parser — extracts structured text from Excel spreadsheets.

Converts each sheet into text: column headers + rows as key-value pairs.
Handles template files (with formulas) by extracting headers and example rows.
"""

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Max rows to extract per sheet (to avoid huge data dumps)
MAX_ROWS = 200


@dataclass
class ParsedSection:
    text: str
    heading: str = ""


def parse_xlsx(file_path: Path) -> list[ParsedSection]:
    """
    Parse an XLSX file into sections (one per sheet).

    Each sheet becomes a text section with column headers and row data.
    """
    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to parse XLSX %s: %s", file_path.name, e)
        return []

    sections: list[ParsedSection] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # First non-empty row as headers
        headers = None
        data_start = 0
        for i, row in enumerate(rows):
            cells = [str(c).strip() if c is not None else "" for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 2:  # At least 2 non-empty cells = header row
                headers = cells
                data_start = i + 1
                break

        if not headers:
            continue

        # Build text representation
        lines = [f"Лист: {sheet_name}"]
        lines.append(f"Колонки: {' | '.join(h for h in headers if h)}")
        lines.append("")

        row_count = 0
        for row in rows[data_start:data_start + MAX_ROWS]:
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Skip completely empty rows
            if not any(cells):
                continue

            # Format as key-value pairs
            pairs = []
            for h, v in zip(headers, cells):
                if h and v:
                    pairs.append(f"{h}: {v}")
            if pairs:
                lines.append("; ".join(pairs))
                row_count += 1

        if row_count > 0:
            sections.append(ParsedSection(
                text="\n".join(lines),
                heading=sheet_name,
            ))

    wb.close()

    total_chars = sum(len(s.text) for s in sections)
    logger.info(
        "Parsed XLSX %s: %d sheets with data, %d chars",
        file_path.name, len(sections), total_chars,
    )
    return sections
