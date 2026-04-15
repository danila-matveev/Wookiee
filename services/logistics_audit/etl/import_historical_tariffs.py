"""
Import historical WB warehouse tariffs from Excel into Supabase.

Usage:
    python -m services.logistics_audit.etl.import_historical_tariffs
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from shared.data_layer._connection import _get_supabase_connection

PROJECT_ROOT = Path(__file__).resolve().parents[3]
load_dotenv(PROJECT_ROOT / ".env")

logger = logging.getLogger(__name__)

DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / "services" / "logistics_audit" / "Тарифы на логискику.xlsx"
DEFAULT_SHEET_NAME = "Тарифы короб"
DEFAULT_BATCH_SIZE = 1000
DEFAULT_PROGRESS_EVERY = 5000

UPSERT_SQL = """
INSERT INTO public.wb_tariffs (
    dt,
    warehouse_name,
    delivery_coef,
    logistics_1l,
    logistics_extra_l,
    storage_1l_day,
    acceptance,
    storage_coef,
    geo_name
)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
ON CONFLICT (dt, warehouse_name) DO UPDATE SET
    delivery_coef = EXCLUDED.delivery_coef,
    logistics_1l = EXCLUDED.logistics_1l,
    logistics_extra_l = EXCLUDED.logistics_extra_l,
    storage_1l_day = EXCLUDED.storage_1l_day,
    acceptance = EXCLUDED.acceptance,
    storage_coef = EXCLUDED.storage_coef,
    geo_name = EXCLUDED.geo_name
"""


@dataclass(frozen=True)
class ImportStats:
    raw_rows: int
    valid_rows: int
    skipped_rows: int
    unique_pairs: int
    duplicate_rows: int
    unique_dates: int
    unique_warehouses: int
    min_date: date | None
    max_date: date | None


def _parse_excel_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _parse_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def load_historical_tariff_rows(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
) -> tuple[list[tuple], ImportStats]:
    """Read workbook rows and map them to wb_tariffs upsert tuples."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[DEFAULT_SHEET_NAME]

    rows: list[tuple] = []
    seen_pairs: set[tuple[date, str]] = set()
    dates: set[date] = set()
    warehouses: set[str] = set()
    raw_rows = 0
    valid_rows = 0
    skipped_rows = 0
    duplicate_rows = 0

    try:
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                continue

            raw_rows += 1
            dt_raw, warehouse_raw, delivery_raw, storage_raw, *_ = row
            dt = _parse_excel_date(dt_raw)
            warehouse_name = str(warehouse_raw).strip() if warehouse_raw is not None else ""
            delivery_coef = _parse_int(delivery_raw)
            storage_coef = _parse_int(storage_raw)

            if dt is None or not warehouse_name or delivery_coef is None or storage_coef is None:
                skipped_rows += 1
                continue

            valid_rows += 1
            pair = (dt, warehouse_name)
            if pair in seen_pairs:
                duplicate_rows += 1
            else:
                seen_pairs.add(pair)

            dates.add(dt)
            warehouses.add(warehouse_name)
            rows.append(
                (
                    dt,
                    warehouse_name,
                    delivery_coef,
                    0,
                    0,
                    0,
                    0,
                    storage_coef,
                    "",
                )
            )
    finally:
        wb.close()

    stats = ImportStats(
        raw_rows=raw_rows,
        valid_rows=valid_rows,
        skipped_rows=skipped_rows,
        unique_pairs=len(seen_pairs),
        duplicate_rows=duplicate_rows,
        unique_dates=len(dates),
        unique_warehouses=len(warehouses),
        min_date=min(dates) if dates else None,
        max_date=max(dates) if dates else None,
    )
    return rows, stats


def import_historical_tariffs(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    batch_size: int = DEFAULT_BATCH_SIZE,
    progress_every: int = DEFAULT_PROGRESS_EVERY,
) -> ImportStats:
    """Upsert workbook history into public.wb_tariffs."""
    rows, stats = load_historical_tariff_rows(workbook_path)
    logger.info(
        "Historical workbook parsed: raw=%s valid=%s skipped=%s unique_pairs=%s duplicates=%s",
        stats.raw_rows,
        stats.valid_rows,
        stats.skipped_rows,
        stats.unique_pairs,
        stats.duplicate_rows,
    )

    if not rows:
        logger.warning("No valid workbook rows found in %s", workbook_path)
        return stats

    conn = _get_supabase_connection()
    try:
        cur = conn.cursor()
        processed = 0
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            cur.executemany(UPSERT_SQL, batch)
            conn.commit()
            processed += len(batch)
            if processed % progress_every == 0 or processed == len(rows):
                logger.info("Historical import progress: %s/%s valid rows", processed, len(rows))
        cur.close()
    finally:
        conn.close()

    logger.info(
        "Historical import complete: valid=%s skipped=%s unique_pairs=%s range=%s..%s",
        stats.valid_rows,
        stats.skipped_rows,
        stats.unique_pairs,
        stats.min_date,
        stats.max_date,
    )
    return stats


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    import_historical_tariffs()


if __name__ == "__main__":
    main()
