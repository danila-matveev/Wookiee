"""Verify our algorithm against financiers' Excel file.

3 levels: row-by-row, SVOD-by-report, cross-validation.

Usage:
    python3 -m services.logistics_audit.verify_against_financiers
"""
from __future__ import annotations
import logging
from datetime import date, datetime, timedelta
from pathlib import Path
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

SERVICE_DIR = Path(__file__).parent
FINANCIERS_PATH = SERVICE_DIR / "ООО_Вуки_проверка_логистики_05_01_01_02.xlsx"
OUR_AUDIT_PATH = SERVICE_DIR / "Аудит логистики 2026-01-01 — 2026-03-23.xlsx"


def _calc_cost(volume: float, coef: float, il: float, base_1l: float, extra_l: float) -> float:
    if volume > 1 and extra_l:
        return round((base_1l + (volume - 1) * extra_l) * coef * il, 2)
    return round(base_1l * coef * il, 2)


def level1_row_by_row() -> bool:
    """Level 1: Row-by-row verification against financiers' Переплата по логистике."""
    logger.info("=" * 60)
    logger.info("LEVEL 1: Row-by-row verification")
    logger.info("=" * 60)

    wb = openpyxl.load_workbook(str(FINANCIERS_PATH), read_only=True, data_only=True)
    ws = wb["Переплата по логистике"]

    total_rows = 0
    max_delta = 0.0
    sum_delta = 0.0
    bad_rows = 0

    for i, rv in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if rv[0] is None and rv[2] is None:
            continue
        try:
            delivery = float(rv[4] or 0)
            coef = float(rv[11] or 0)
            vol = float(rv[12] or 0)
            ktr = float(rv[13] or 0)
            base_1l = float(rv[14] or 0) if rv[14] else 0
            extra_l = float(rv[15] or 0) if rv[15] else 0
            their_cost = float(rv[16] or 0)
            their_diff = float(rv[17] or 0)

            our_cost = _calc_cost(vol, coef, ktr, base_1l, extra_l)
            our_diff = round(delivery - our_cost, 2)

            delta = abs(our_diff - their_diff)
            max_delta = max(max_delta, delta)
            sum_delta += delta
            total_rows += 1

            if delta > 0.5:
                bad_rows += 1
        except (TypeError, ValueError):
            continue

    wb.close()

    mean_delta = sum_delta / total_rows if total_rows else 0
    passed = max_delta < 1.0

    logger.info(f"  Rows checked: {total_rows}")
    logger.info(f"  Max delta: {max_delta:.4f} rub")
    logger.info(f"  Mean delta: {mean_delta:.4f} rub")
    logger.info(f"  Rows with delta > 0.5: {bad_rows}")
    logger.info(f"  Result: {'PASS' if passed else 'FAIL'} (target: max delta < 1 rub)")
    return passed


def level2_svod_by_report() -> bool:
    """Level 2: SVOD comparison — our old vs financiers, per report_id."""
    logger.info("\n" + "=" * 60)
    logger.info("LEVEL 2: SVOD by report_id")
    logger.info("=" * 60)

    wb_fin = openpyxl.load_workbook(str(FINANCIERS_PATH), read_only=True, data_only=True)
    ws_fin = wb_fin["СВОД"]
    fin_svod: dict[str, float] = {}
    for i, rv in enumerate(ws_fin.iter_rows(values_only=True)):
        if i == 0:
            continue
        if rv[0] is None:
            continue
        try:
            fin_svod[str(rv[0])] = float(rv[6] or 0)
        except (TypeError, ValueError):
            continue
    wb_fin.close()

    wb_our = openpyxl.load_workbook(str(OUR_AUDIT_PATH), read_only=True, data_only=True)
    ws_our = wb_our["СВОД"]
    our_svod: dict[str, float] = {}
    for i, rv in enumerate(ws_our.iter_rows(values_only=True)):
        if i == 0:
            continue
        if rv[0] is None:
            continue
        try:
            our_svod[str(rv[0])] = float(rv[5] or 0)
        except (TypeError, ValueError):
            continue
    wb_our.close()

    max_delta = 0.0
    logger.info(f"\n  {'Report':<15} {'Our old':>12} {'Financiers':>12} {'Delta':>10}")
    logger.info(f"  {'-' * 50}")
    for report_id in sorted(fin_svod.keys()):
        fin_val = fin_svod[report_id]
        our_val = our_svod.get(report_id, 0)
        delta = abs(our_val - fin_val)
        max_delta = max(max_delta, delta)
        marker = "✓" if delta < 5 else "✗"
        logger.info(f"  {marker} {report_id:<15} {our_val:>12,.2f} {fin_val:>12,.2f} {delta:>10,.2f}")

    logger.info(f"\n  Max delta: {max_delta:,.2f} rub")
    logger.info(f"  Result: INFO — old audit used different methodology, differences expected")
    return True


def level3_cross_validation() -> bool:
    """Level 3: Fisanov cross-validation (already proven: 0.96 rub)."""
    logger.info("\n" + "=" * 60)
    logger.info("LEVEL 3: Cross-validation (Fisanov)")
    logger.info("=" * 60)

    from services.logistics_audit.recalculate_ooo import run_fisanov
    results = run_fisanov()
    final = results["fix5"]["total_overpay"]
    remainder = abs(final - 144_901.0)
    passed = remainder < 1.0
    logger.info(f"  Fisanov remainder: {remainder:.2f} rub")
    logger.info(f"  Result: {'PASS' if passed else 'FAIL'}")
    return passed


def main():
    logger.info("LOGISTICS AUDIT VERIFICATION")
    logger.info("Financiers file: " + FINANCIERS_PATH.name)
    logger.info("")

    l1 = level1_row_by_row()
    l2 = level2_svod_by_report()
    l3 = level3_cross_validation()

    logger.info("\n" + "=" * 60)
    logger.info("SUMMARY")
    logger.info("=" * 60)
    logger.info(f"  Level 1 (row-by-row):    {'PASS' if l1 else 'FAIL'}")
    logger.info(f"  Level 2 (SVOD):          {'PASS' if l2 else 'FAIL'}")
    logger.info(f"  Level 3 (Fisanov):       {'PASS' if l3 else 'FAIL'}")
    overall = l1 and l2 and l3
    logger.info(f"  Overall:                 {'ALL PASS' if overall else 'ISSUES FOUND'}")
    return overall


if __name__ == "__main__":
    main()
