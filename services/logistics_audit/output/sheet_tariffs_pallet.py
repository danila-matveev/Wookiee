"""Sheet 11: 'Тариф монопалета' — pallet tariff data."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet

def write_tariffs_pallet(ws: Worksheet, pallet_data: dict) -> None:
    """Write raw pallet tariff data. Structure TBD based on API response."""
    warehouses = pallet_data.get("response", {}).get("data", {}).get("warehouseList", [])
    if not warehouses:
        ws.cell(1, 1, "Нет данных по тарифам монопалета")
        return
    headers = list(warehouses[0].keys())
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    for i, wh in enumerate(warehouses, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(i, col, wh.get(h, ""))
