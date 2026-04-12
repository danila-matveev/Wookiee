"""Sheet 9: 'Габариты в карточке' — card dimensions."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet

HEADERS = ["Код номенклатуры", "Ширина (см)", "Высота (см)", "Длина (см)", "Объём (л)"]

def write_dimensions(ws: Worksheet, card_dims: dict[int, dict]) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)
    for i, (nm_id, dims) in enumerate(sorted(card_dims.items()), 2):
        ws.cell(i, 1, nm_id)
        ws.cell(i, 2, dims.get("width", 0))
        ws.cell(i, 3, dims.get("height", 0))
        ws.cell(i, 4, dims.get("length", 0))
        ws.cell(i, 5, dims.get("volume", 0))
