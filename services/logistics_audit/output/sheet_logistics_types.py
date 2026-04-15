"""Sheet 7: 'Виды логистики' — GROUP BY bonus_type_name."""
from __future__ import annotations
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.report_row import ReportRow

HEADERS = ["Тип логистики", "Кол-во", "Сумма", "Средняя"]

def write_logistics_types(ws: Worksheet, rows: list[ReportRow]) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    agg: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        if row.is_logistics:
            agg[row.bonus_type_name].append(row.delivery_rub)

    sorted_agg = sorted(agg.items(), key=lambda x: sum(x[1]), reverse=True)
    for i, (btype, amounts) in enumerate(sorted_agg, 2):
        total = sum(amounts)
        ws.cell(i, 1, btype)
        ws.cell(i, 2, len(amounts))
        ws.cell(i, 3, round(total, 2))
        ws.cell(i, 4, round(total / len(amounts), 2) if amounts else 0)
