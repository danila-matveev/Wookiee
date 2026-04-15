"""Sheet 6: 'Переплата по артикулам' — GROUP BY nm_id."""
from __future__ import annotations
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.calculators.logistics_overpayment import OverpaymentResult
from services.logistics_audit.models.report_row import ReportRow

HEADERS = ["Код номенклатуры", "Кол-во строк", "Сумма переплаты", "Средняя переплата"]

def write_pivot_by_article(
    ws: Worksheet,
    rows: list[ReportRow],
    results: list[OverpaymentResult | None],
) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    agg: dict[int, list[float]] = defaultdict(list)
    for row, res in zip(rows, results):
        if res is not None:
            agg[row.nm_id].append(res.overpayment)

    sorted_agg = sorted(agg.items(), key=lambda x: sum(x[1]), reverse=True)
    for i, (nm_id, overpays) in enumerate(sorted_agg, 2):
        total = sum(overpays)
        ws.cell(i, 1, nm_id)
        ws.cell(i, 2, len(overpays))
        ws.cell(i, 3, round(total, 2))
        ws.cell(i, 4, round(total / len(overpays), 2) if overpays else 0)
