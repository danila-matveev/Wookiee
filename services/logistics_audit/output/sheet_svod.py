"""Sheet 3: 'СВОД' — summary by realizationreport_id."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.report_row import ReportRow

HEADERS = [
    "№ отчёта", "Юр. лицо", "Дата начала", "Дата конца",
    "Стоимость логистики", "Переплата", "Коррекция логистики",
    "Сторно логистики", "Расчётная стоимость",
]


def write_svod(
    ws: Worksheet,
    all_rows: list[ReportRow],
    overpayments_by_report: dict[int, float],
) -> None:
    """Write SVOD sheet: one row per realizationreport_id."""
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    # Aggregate by report
    reports: dict[int, dict] = {}
    for row in all_rows:
        rid = row.realizationreport_id
        if rid not in reports:
            reports[rid] = {
                "supplier_name": row.ppvz_supplier_name,
                "date_from": row.date_from,
                "date_to": row.date_to,
                "logistics": 0.0,
                "correction": 0.0,
            }
        if row.supplier_oper_name == "Логистика":
            reports[rid]["logistics"] += row.delivery_rub
        elif row.supplier_oper_name == "Коррекция логистики":
            reports[rid]["correction"] += row.delivery_rub

    for i, (rid, info) in enumerate(sorted(reports.items()), 2):
        overpay = overpayments_by_report.get(rid, 0)
        ws.cell(i, 1, rid)
        ws.cell(i, 2, info["supplier_name"])
        ws.cell(i, 3, info["date_from"])
        ws.cell(i, 4, info["date_to"])
        ws.cell(i, 5, round(info["logistics"], 2))
        ws.cell(i, 6, round(overpay, 2))
        ws.cell(i, 7, round(info["correction"], 2))
        ws.cell(i, 8, 0)  # Сторно — из данных
        ws.cell(i, 9, round(info["logistics"] - overpay + info["correction"], 2))
