"""Sheet 2: 'Переплата по логистике' — pre-calculated values, no formulas."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.report_row import ReportRow
from services.logistics_audit.calculators.logistics_overpayment import OverpaymentResult

HEADERS = [
    "№ отчёта", "Номер поставки", "Код номенклатуры", "Дата заказа",
    "Услуги по доставке", "Склад", "ШК", "Srid", "Фикс. коэф.",
    "Коэф. для расчёта", "Объём", "КТР", "Расчётная стоимость",
    "Разница (переплата)",
]


def write_overpayment_values(
    ws: Worksheet,
    rows: list[ReportRow],
    results: list[OverpaymentResult | None],
    volumes: dict[int, float],
    coefs: list[float],
    row_ils: list[float] | None = None,
) -> None:
    """Write Sheet 2 with pre-calculated overpayment values.

    Per lawyer recommendation: negative differences are completely excluded
    from this sheet (not just marked).
    """
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    total_overpayment = 0.0
    excel_row = 2
    for i, (row, res, coef) in enumerate(zip(rows, results, coefs)):
        if res is None:
            continue
        # Exclude negative differences entirely (lawyer recommendation №5)
        if res.overpayment < 0:
            continue

        ws.cell(excel_row, 1, row.realizationreport_id)
        ws.cell(excel_row, 2, row.gi_id)
        ws.cell(excel_row, 3, row.nm_id)
        ws.cell(excel_row, 4, str(row.order_dt) if row.order_dt else "")
        ws.cell(excel_row, 5, row.delivery_rub)
        ws.cell(excel_row, 6, row.office_name)
        ws.cell(excel_row, 7, row.shk_id)
        ws.cell(excel_row, 8, row.srid)
        ws.cell(excel_row, 9, row.dlv_prc)
        ws.cell(excel_row, 10, coef)
        ws.cell(excel_row, 11, volumes.get(row.nm_id, 0))
        if row_ils is not None and i < len(row_ils):
            ws.cell(excel_row, 12, row_ils[i])
        else:
            ws.cell(excel_row, 12, res.calculated_cost)
        ws.cell(excel_row, 13, res.calculated_cost)
        ws.cell(excel_row, 14, res.overpayment)
        total_overpayment += res.overpayment
        excel_row += 1

    # Summary row at top
    ws.insert_rows(1)
    ws.cell(1, 1, "ИТОГО переплата:")
    ws.cell(1, 14, round(total_overpayment, 2))
