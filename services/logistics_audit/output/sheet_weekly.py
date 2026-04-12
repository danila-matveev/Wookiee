"""Sheet 8: 'Еженед. отчет' — GROUP BY realizationreport_id."""
from __future__ import annotations
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.report_row import ReportRow

HEADERS = [
    "№ отчёта", "Розничная сумма продаж", "К перечислению",
    "Логистика", "Повышенная логистика", "Штрафы",
    "Хранение", "Приёмка", "Удержания",
]

def write_weekly(ws: Worksheet, all_rows: list[ReportRow]) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    agg: dict[int, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in all_rows:
        rid = row.realizationreport_id
        if row.doc_type_name == "Продажа":
            agg[rid]["retail"] += row.retail_amount
        agg[rid]["ppvz"] += row.ppvz_for_pay
        agg[rid]["delivery"] += row.delivery_rub
        agg[rid]["rebill"] += row.rebill_logistic_cost
        agg[rid]["penalty"] += row.penalty
        agg[rid]["storage"] += row.storage_fee
        agg[rid]["acceptance"] += row.acceptance
        agg[rid]["deduction"] += row.deduction

    for i, (rid, vals) in enumerate(sorted(agg.items()), 2):
        ws.cell(i, 1, rid)
        ws.cell(i, 2, round(vals["retail"], 2))
        ws.cell(i, 3, round(vals["ppvz"], 2))
        ws.cell(i, 4, round(vals["delivery"], 2))
        ws.cell(i, 5, round(vals["rebill"], 2))
        ws.cell(i, 6, round(vals["penalty"], 2))
        ws.cell(i, 7, round(vals["storage"], 2))
        ws.cell(i, 8, round(vals["acceptance"], 2))
        ws.cell(i, 9, round(vals["deduction"], 2))
