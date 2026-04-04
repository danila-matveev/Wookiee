"""Sheet 1: 'Переплата по логистике (короб)' — with live Excel formulas."""
from __future__ import annotations
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.report_row import ReportRow

HEADERS = [
    "Номер поставки",      # A: gi_id
    "Код номенклатуры",     # B: nm_id
    "Дата заказа",          # C: order_dt
    "",                     # D: unused
    "Услуги по доставке",   # E: delivery_rub
    "Дата начала фиксации", # F: fix_tariff_date_from
    "Дата конца фиксации",  # G: fix_tariff_date_to
    "Склад",                # H: office_name
    "ШК",                   # I: shk_id
    "Srid",                 # J: srid
    "Фикс. коэф.",         # K: dlv_prc
    "Коэф. для расчёта",   # L: formula
    "Объём из карточки",    # M: VLOOKUP
    "Объём из остатков",    # N: warehouse_remains
    "КТР",                  # O: ktr
    "Стоимость 1л",         # P: base_1l
    "Стоимость доп.л",      # Q: extra_l
    "Стоимость логистики",  # R: formula
    "Разница",              # S: formula
    "Включено в итог",      # T: formula
]


def write_overpayment_formulas(
    ws: Worksheet,
    rows: list[ReportRow],
    ktr: float,
    base_1l: float,
    extra_l: float,
    row_ils: list[float] | None = None,
) -> None:
    """Write Sheet 1 with live Excel formulas.

    If row_ils is provided, each row gets its actual weekly IL.
    Otherwise falls back to static ktr.
    """
    # Header
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)

    for i, row in enumerate(rows, 2):
        idx = i - 2  # index into row_ils
        ws.cell(i, 1, row.gi_id)
        ws.cell(i, 2, row.nm_id)
        ws.cell(i, 3, str(row.order_dt) if row.order_dt else "")
        # D is unused
        ws.cell(i, 5, row.delivery_rub)
        ws.cell(i, 6, str(row.fix_tariff_date_from) if row.fix_tariff_date_from else "")
        ws.cell(i, 7, str(row.fix_tariff_date_to) if row.fix_tariff_date_to else "")
        ws.cell(i, 8, row.office_name)
        ws.cell(i, 9, row.shk_id)
        ws.cell(i, 10, row.srid)
        ws.cell(i, 11, row.dlv_prc)
        # L: Coefficient for calculation
        ws.cell(i, 12, f'=IF(K{i}>0,K{i},VLOOKUP(H{i},\'Тарифы короб\'!A:C,3,FALSE)/100)')
        # M: Volume from card lookup
        ws.cell(i, 13, f'=IFERROR(VLOOKUP(B{i},\'Габариты в карточке\'!A:E,5,FALSE),"")')
        # N: volume from remains (filled later by excel_generator)
        ws.cell(i, 14, "")
        # O: КТР — per-row weekly IL if available
        if row_ils is not None and idx < len(row_ils):
            ws.cell(i, 15, row_ils[idx])
        else:
            ws.cell(i, 15, ktr)
        ws.cell(i, 16, base_1l)
        ws.cell(i, 17, extra_l)
        # R: logistics cost formula
        ws.cell(i, 18, f'=IF(M{i}>1,(P{i}+(M{i}-1)*Q{i})*L{i}*O{i},P{i}*L{i}*O{i})')
        # S: difference
        ws.cell(i, 19, f'=E{i}-R{i}')
        # T: included in total (only positive differences)
        ws.cell(i, 20, f'=IF(S{i}>=0,"Да","Нет")')
