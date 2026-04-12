"""Sheet 10: 'Тарифы короб' — box tariff snapshot."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet
from services.logistics_audit.models.tariff_snapshot import TariffSnapshot

HEADERS = ["Склад", "Регион", "Коэф. логистики (%)", "Коэф. хранения (%)",
           "Логистика: база (₽)", "Логистика: доп.л (₽)", "Хранение: база (₽)"]

def write_tariffs_box(ws: Worksheet, tariffs: dict[str, TariffSnapshot]) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)
    for i, (name, t) in enumerate(sorted(tariffs.items()), 2):
        ws.cell(i, 1, name)
        ws.cell(i, 2, t.geo_name)
        ws.cell(i, 3, t.delivery_coef_pct)
        ws.cell(i, 4, t.storage_coef_pct)
        ws.cell(i, 5, t.box_delivery_base)
        ws.cell(i, 6, t.box_delivery_liter)
        ws.cell(i, 7, t.box_storage_base)
