"""Sheet 5: 'ИЛ' — localization index."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet

HEADERS = ["Дата обновления", "ИЛ", "Дата начала действия", "Дата конца действия"]

def write_il(ws: Worksheet, il_data: list[dict] | None = None) -> None:
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)
    if il_data:
        for i, entry in enumerate(il_data, 2):
            ws.cell(i, 1, entry.get("date", ""))
            ws.cell(i, 2, entry.get("il", ""))
            ws.cell(i, 3, entry.get("date_from", ""))
            ws.cell(i, 4, entry.get("date_to", ""))
