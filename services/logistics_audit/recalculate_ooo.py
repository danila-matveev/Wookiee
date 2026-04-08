"""Recalculate OOO logistics audit with 5 cumulative fixes.

Reads local Excel files (no WB API needed), applies fixes one-by-one,
decomposes the discrepancy, and generates an output Excel.

Usage:
    python3 -m services.logistics_audit.recalculate_ooo           # OOO recalculation
    python3 -m services.logistics_audit.recalculate_ooo --fisanov  # Validate on Fisanov
"""
from __future__ import annotations

import argparse
import logging
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl

from services.logistics_audit.calculators.tariff_periods import get_base_tariffs
from services.logistics_audit.models.report_row import FORWARD_DELIVERY_TYPES

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

SERVICE_DIR = Path(__file__).parent

# --- File paths ---
OOO_AUDIT_PATH = SERVICE_DIR / "Аудит логистики 2026-01-01 — 2026-03-23.xlsx"
TARIFF_FILE_PATH = SERVICE_DIR / "Тарифы на логискику.xlsx"
FISANOV_PATH = SERVICE_DIR / "ИП Фисанов. Проверка логистики 05.01.2026 г. - 01.02.2026 г._Итоговый.xlsx"

# --- OOO IL values from WB dashboard (financiers' verified file) ---
OOO_IL_DASHBOARD: dict[date, float] = {
    date(2025, 11, 24): 1.04,
    date(2025, 12, 1): 1.04,
    date(2025, 12, 8): 1.04,
    date(2025, 12, 15): 1.04,
    date(2025, 12, 22): 1.11,
    date(2025, 12, 29): 1.10,
    date(2026, 1, 5): 1.09,
    date(2026, 1, 12): 1.09,
    date(2026, 1, 19): 1.08,
    date(2026, 1, 26): 1.07,
    date(2026, 2, 2): 1.06,
    date(2026, 2, 9): 1.06,
    date(2026, 2, 16): 1.05,
    date(2026, 2, 23): 1.04,
    date(2026, 3, 2): 1.01,
    date(2026, 3, 9): 0.99,
    date(2026, 3, 16): 1.01,
    date(2026, 3, 23): 0.98,
}

# Fisanov IL values (from validate_fisanov.py)
FISANOV_IL: dict[date, float] = {
    date(2025, 12, 1): 1.29,
    date(2025, 12, 8): 1.33,
    date(2025, 12, 15): 1.34,
    date(2025, 12, 22): 1.37,
    date(2025, 12, 29): 1.37,
    date(2026, 1, 5): 1.36,
    date(2026, 1, 12): 1.34,
    date(2026, 1, 19): 1.35,
    date(2026, 1, 26): 1.33,
}


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None
    if pd.notna(v) if isinstance(v, float) else v is not None:
        try:
            return date.fromisoformat(str(v)[:10])
        except (ValueError, TypeError):
            return None
    return None


# ─── Data loaders ────────────────────────────────────────────────────────


def load_tariff_file(path: Path = TARIFF_FILE_PATH) -> dict[str, list[tuple[date, float]]]:
    """Load warehouse coefficients from Тарифы на логискику.xlsx.

    Returns pre-indexed: {warehouse_name: [(date, coef_decimal), ...]} sorted newest-first.
    """
    logger.info(f"Loading tariff file: {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Тарифы короб"]

    index: dict[str, list[tuple[date, float]]] = {}
    count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        dt_raw, warehouse, coef_pct, *_ = row
        dt = _to_date(dt_raw)
        if dt is None or warehouse is None or coef_pct is None:
            continue
        try:
            wh = str(warehouse).strip()
            coef = float(coef_pct) / 100.0
            if wh not in index:
                index[wh] = []
            index[wh].append((dt, coef))
            count += 1
        except (ValueError, TypeError):
            continue

    wb.close()
    for wh in index:
        index[wh].sort(reverse=True)
    logger.info(f"  Loaded {count} tariff entries for {len(index)} warehouses")
    return index


def _find_tariff_coef(
    wh_index: dict[str, list[tuple[date, float]]],
    order_date: date,
    warehouse: str,
) -> float | None:
    """Fast coefficient lookup using pre-indexed tariffs."""
    entries = wh_index.get(warehouse)
    if not entries:
        return None
    for dt, coef in entries:
        if dt <= order_date:
            return coef
    return None


def _calc_cost(volume: float, coef: float, il: float, base_1l: float, extra_l: float) -> float:
    """Calculate expected logistics cost."""
    if volume > 1:
        base = (base_1l + (volume - 1) * extra_l) * coef
    else:
        base = base_1l * coef
    return round(base * il, 2)


# ─── OOO: Load from original Переплата sheet + join with Детализация ─────


def load_ooo_data(
    path: Path = OOO_AUDIT_PATH,
) -> list[dict]:
    """Load OOO audit rows from 'Переплата по логистике' sheet,
    joined with 'Детализация' for bonus_type_name and fixation dates."""
    logger.info(f"Loading original audit: {path.name}")

    # 1. Load Переплата sheet (original per-row calculations)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Переплата по логистике"]
    pereplata_rows = []
    for i, rv in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # ИТОГО row
            continue
        if i == 1:  # headers
            continue
        if rv[0] is None and rv[2] is None:
            continue
        try:
            pereplata_rows.append({
                "report_id": rv[0],
                "nm_id": int(rv[2] or 0),
                "order_dt": _to_date(rv[3]),
                "delivery_rub": float(rv[4] or 0),
                "warehouse": str(rv[5] or "").strip(),
                "shk_id": rv[6],
                "srid": str(rv[7] or ""),
                "fixed_coef": float(rv[8] or 0),
                "calc_coef": float(rv[9] or 0),
                "volume": float(rv[10] or 0),
                "ktr_old": float(rv[11] or 0),
                "calc_cost_old": float(rv[12] or 0),
                "diff_old": float(rv[13] or 0),
            })
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    logger.info(f"  Переплата rows: {len(pereplata_rows):,}")
    logger.info(f"  Original total: {sum(r['diff_old'] for r in pereplata_rows):,.2f} rub")

    # 2. Load Детализация for bonus_type_name and fixation dates
    logger.info("  Loading Детализация for join...")
    df_det = pd.read_excel(str(path), sheet_name="Детализация", engine="openpyxl")
    df_log = df_det[df_det["supplier_oper_name"] == "Логистика"].copy()

    srid_lookup: dict[str, list[dict]] = defaultdict(list)
    for _, r in df_log.iterrows():
        srid_lookup[str(r["srid"])].append({
            "bonus_type": str(r.get("bonus_type_name", "")),
            "fix_from": _to_date(r.get("fix_tariff_date_from")),
            "fix_to": _to_date(r.get("fix_tariff_date_to")),
            "delivery_rub": float(r.get("delivery_rub", 0)),
            "dlv_prc": float(r.get("dlv_prc", 0)),
        })

    # 3. Join: enrich each Переплата row with bonus_type_name
    joined = 0
    for row in pereplata_rows:
        candidates = srid_lookup.get(row["srid"], [])
        # Match by delivery_rub for disambiguation
        best = None
        for c in candidates:
            if abs(c["delivery_rub"] - row["delivery_rub"]) < 0.01:
                best = c
                break
        if best is None and candidates:
            best = candidates[0]
        if best:
            row["bonus_type"] = best["bonus_type"]
            row["fix_from"] = best["fix_from"]
            row["fix_to"] = best["fix_to"]
            row["dlv_prc"] = best["dlv_prc"]
            joined += 1
        else:
            row["bonus_type"] = "К клиенту при продаже"  # assume forward
            row["fix_from"] = None
            row["fix_to"] = None
            row["dlv_prc"] = row["calc_coef"]

    logger.info(f"  Joined: {joined:,} / {len(pereplata_rows):,}")

    # Log bonus_type distribution
    from collections import Counter
    bt_counts = Counter(r["bonus_type"] for r in pereplata_rows)
    for bt, cnt in bt_counts.most_common():
        fwd = "✅" if bt in FORWARD_DELIVERY_TYPES else "❌"
        logger.info(f"    {fwd} {bt}: {cnt:,}")

    return pereplata_rows


def load_fisanov_rows() -> list[dict]:
    """Load Fisanov reference rows from 'Переплата по логистике' sheet."""
    logger.info(f"Loading Fisanov from: {FISANOV_PATH.name}")
    wb = openpyxl.load_workbook(str(FISANOV_PATH), read_only=True, data_only=True)
    ws = wb["Переплата по логистике"]

    rows = []
    for i, rv in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if rv[0] is None and rv[2] is None:
            continue
        try:
            rows.append({
                "nm_id": int(rv[2]) if rv[2] else 0,
                "order_dt": _to_date(rv[3]),
                "delivery_rub": float(rv[4] or 0),
                "fix_from": _to_date(rv[5]),
                "fix_to": _to_date(rv[6]),
                "warehouse": str(rv[7] or ""),
                "shk_id": rv[8],
                "srid": rv[9],
                "fixed_coef": float(rv[10] or 0),
                "calc_coef": float(rv[11] or 0),
                "volume": float(rv[12] or 0),
                "ktr_old": float(rv[13] or 0),
                "calc_cost_old": float(rv[14] or 0) if len(rv) > 14 else 0,
                "diff_old": float(rv[15] or 0) if len(rv) > 15 else 0,
                "bonus_type": "К клиенту при продаже",  # all Fisanov = forward
                "dlv_prc": float(rv[11] or 0),  # calc_coef = coef used
            })
        except (TypeError, ValueError, IndexError):
            continue

    wb.close()
    logger.info(f"  Loaded {len(rows)} rows")
    return rows


# ─── Calculation engine ──────────────────────────────────────────────────


def _resolve_coef_3tier(
    row: dict,
    tariffs_index: dict[str, list[tuple[date, float]]],
) -> tuple[float, str]:
    """Resolve warehouse coefficient with 3-tier priority from spec."""
    order_dt = row["order_dt"]
    dlv_prc = row.get("dlv_prc", 0) or row.get("calc_coef", 0)
    fix_to = row.get("fix_to")

    # Tier 1: Fixation active (fix_to > order_dt)
    if dlv_prc > 0 and fix_to and order_dt and fix_to > order_dt:
        return dlv_prc, "fixation"

    # Tier 2: Fixation expired or no fixation → tariff file
    if order_dt:
        file_coef = _find_tariff_coef(tariffs_index, order_dt, row["warehouse"])
        if file_coef is not None and file_coef > 0:
            return file_coef, "tariff_file"

    # Tier 3: fallback to dlv_prc/calc_coef
    if dlv_prc > 0:
        return dlv_prc, "fallback"

    return 0.0, "none"


def recalculate_ooo(
    rows: list[dict],
    tariffs_index: dict[str, list[tuple[date, float]]],
    il_dashboard: dict[date, float],
) -> dict[str, dict]:
    """Run cumulative fix decomposition for OOO data.

    Uses original Переплата sheet as baseline (sum of diff_old = 174,499).
    """
    results: dict[str, dict] = {}

    # ── Baseline: original audit values from sheet ──
    baseline_total = sum(r["diff_old"] for r in rows)
    results["baseline"] = {
        "total_overpay": round(baseline_total, 2),
        "rows_count": len(rows),
        "label": "Baseline (original audit)",
    }

    # ── +Fix1: Remove non-forward delivery rows ──
    forward_rows = [r for r in rows if r["bonus_type"] in FORWARD_DELIVERY_TYPES]
    excluded_rows = [r for r in rows if r["bonus_type"] not in FORWARD_DELIVERY_TYPES]
    fix1_excluded_overpay = sum(r["diff_old"] for r in excluded_rows)
    fix1_total = baseline_total - fix1_excluded_overpay

    results["fix1"] = {
        "total_overpay": round(fix1_total, 2),
        "rows_count": len(forward_rows),
        "rows_excluded": len(excluded_rows),
        "excluded_overpay": round(fix1_excluded_overpay, 2),
        "label": "+Fix1: Forward delivery filter",
    }

    # ── +Fix2: Period-based tariffs (replace old calibrated tariffs) ──
    # Recalculate each forward row with period-based tariffs + old coef + old IL
    fix2_total = 0.0
    fix2_count = 0
    for r in forward_rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        cost = _calc_cost(r["volume"], r["calc_coef"], r["ktr_old"], base_1l, extra_l)
        overpay = round(r["delivery_rub"] - cost, 2)
        r["fix2_cost"] = cost
        r["fix2_overpay"] = overpay
        fix2_total += overpay
        fix2_count += 1

    results["fix2"] = {
        "total_overpay": round(fix2_total, 2),
        "rows_count": fix2_count,
        "label": "+Fix1+2: Period-based tariffs",
    }

    # ── +Fix3: Warehouse coef from tariff file ──
    fix3_total = 0.0
    fix3_count = 0
    coef_sources: dict[str, int] = defaultdict(int)
    for r in forward_rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        coef, src = _resolve_coef_3tier(r, tariffs_index)
        coef_sources[src] += 1
        if coef == 0:
            r["fix3_cost"] = r["delivery_rub"]
            r["fix3_overpay"] = 0
            continue

        cost = _calc_cost(r["volume"], coef, r["ktr_old"], base_1l, extra_l)
        overpay = round(r["delivery_rub"] - cost, 2)
        r["fix3_cost"] = cost
        r["fix3_overpay"] = overpay
        r["fix3_coef"] = coef
        r["fix3_coef_src"] = src
        fix3_total += overpay
        fix3_count += 1

    results["fix3"] = {
        "total_overpay": round(fix3_total, 2),
        "rows_count": fix3_count,
        "label": "+Fix1+2+3: Warehouse coef from tariff file",
        "coef_sources": dict(coef_sources),
    }

    # ── +Fix4: IL from WB dashboard ──
    fix4_total = 0.0
    fix4_count = 0
    fix4_details: list[dict] = []
    for r in forward_rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        coef = r.get("fix3_coef", r["calc_coef"])
        if coef == 0:
            continue

        il = 1.0
        if r["order_dt"]:
            mon = _monday(r["order_dt"])
            il = il_dashboard.get(mon, 1.0)

        cost = _calc_cost(r["volume"], coef, il, base_1l, extra_l)
        overpay = round(r["delivery_rub"] - cost, 2)
        r["fix4_cost"] = cost
        r["fix4_overpay"] = overpay
        fix4_total += overpay
        fix4_count += 1
        fix4_details.append({
            "report_id": r.get("report_id"),
            "gi_id": r.get("gi_id"),
            "nm_id": r["nm_id"],
            "order_dt": r["order_dt"],
            "delivery_rub": r["delivery_rub"],
            "fix_from": r.get("fix_from"),
            "fix_to": r.get("fix_to"),
            "warehouse": r["warehouse"],
            "shk_id": r.get("shk_id"),
            "srid": r.get("srid"),
            "fixed_coef": r.get("fixed_coef", 0),
            "volume": r["volume"],
            "coef_old": r["calc_coef"],
            "coef_new": coef,
            "coef_src": r.get("fix3_coef_src", "original"),
            "il_old": r["ktr_old"],
            "il_new": il,
            "base_1l": base_1l,
            "extra_l": extra_l,
            "calc_cost_old": r["calc_cost_old"],
            "calc_cost_new": cost,
            "overpay_old": r["diff_old"],
            "overpay_new": overpay,
            "bonus_type": r["bonus_type"],
        })

    results["fix4"] = {
        "total_overpay": round(fix4_total, 2),
        "rows_count": fix4_count,
        "label": "+Fix1+2+3+4: IL from WB dashboard",
        "details": fix4_details,
    }

    # ── +Fix5: Exclude negative overpayments ──
    fix5_positive = 0.0
    fix5_negative = 0.0
    fix5_pos_count = 0
    fix5_neg_count = 0
    for d in fix4_details:
        if d["overpay_new"] >= 0:
            fix5_positive += d["overpay_new"]
            fix5_pos_count += 1
        else:
            fix5_negative += d["overpay_new"]
            fix5_neg_count += 1

    results["fix5"] = {
        "total_overpay": round(fix5_positive, 2),
        "rows_count": fix5_pos_count,
        "rows_excluded": fix5_neg_count,
        "negative_total": round(fix5_negative, 2),
        "label": "+Fix1+2+3+4+5: Exclude negatives",
        "details": [d for d in fix4_details if d["overpay_new"] >= 0],
    }

    return results


def recalculate_fisanov(
    rows: list[dict],
    tariffs_index: dict[str, list[tuple[date, float]]],
    il_table: dict[date, float],
) -> dict[str, dict]:
    """Run fix decomposition for Fisanov validation."""
    results: dict[str, dict] = {}

    # Baseline: old logic (tariff=46/14, coef=calc_coef, KTR=1.0)
    baseline_total = 0.0
    for r in rows:
        cost = _calc_cost(r["volume"], r["calc_coef"], 1.0, 46.0, 14.0)
        overpay = r["delivery_rub"] - cost
        baseline_total += overpay
    results["baseline"] = {
        "total_overpay": round(baseline_total, 2),
        "rows_count": len(rows),
        "label": "Baseline (old: tariff=46, KTR=1.0)",
    }

    # Fix1: All Fisanov rows are forward → delta = 0
    results["fix1"] = {
        "total_overpay": round(baseline_total, 2),
        "rows_count": len(rows),
        "rows_excluded": 0,
        "label": "+Fix1: Forward delivery filter",
    }

    # Fix2: Period-based tariffs
    fix2_total = 0.0
    for r in rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        cost = _calc_cost(r["volume"], r["calc_coef"], 1.0, base_1l, extra_l)
        overpay = r["delivery_rub"] - cost
        r["fix2_overpay"] = overpay
        fix2_total += overpay
    results["fix2"] = {
        "total_overpay": round(fix2_total, 2),
        "rows_count": len(rows),
        "label": "+Fix1+2: Period-based tariffs",
    }

    # Fix3: Warehouse coef (Fisanov has fixed_coef in data)
    fix3_total = 0.0
    for r in rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        coef = r["calc_coef"]
        if r["fixed_coef"] > 0 and r["fix_to"] and r["order_dt"] and r["fix_to"] > r["order_dt"]:
            coef = r["fixed_coef"]
        cost = _calc_cost(r["volume"], coef, 1.0, base_1l, extra_l)
        overpay = r["delivery_rub"] - cost
        r["fix3_coef"] = coef
        r["fix3_overpay"] = overpay
        fix3_total += overpay
    results["fix3"] = {
        "total_overpay": round(fix3_total, 2),
        "rows_count": len(rows),
        "label": "+Fix1+2+3: Warehouse coef",
    }

    # Fix4: IL from dashboard
    fix4_total = 0.0
    fix4_details: list[dict] = []
    for r in rows:
        base_1l, extra_l = get_base_tariffs(
            r["order_dt"], r["fix_from"], r["fix_to"], r["volume"],
        )
        coef = r.get("fix3_coef", r["calc_coef"])
        il = 1.0
        if r["order_dt"]:
            mon = _monday(r["order_dt"])
            il = il_table.get(mon, r.get("ktr_old", 1.0))
        cost = _calc_cost(r["volume"], coef, il, base_1l, extra_l)
        overpay = round(r["delivery_rub"] - cost, 2)
        fix4_total += overpay
        fix4_details.append({
            "nm_id": r["nm_id"],
            "order_dt": r["order_dt"],
            "delivery_rub": r["delivery_rub"],
            "volume": r["volume"],
            "coef_new": coef,
            "il_new": il,
            "base_1l": base_1l,
            "extra_l": extra_l,
            "calc_cost_new": cost,
            "overpay_new": overpay,
            "warehouse": r["warehouse"],
        })
    results["fix4"] = {
        "total_overpay": round(fix4_total, 2),
        "rows_count": len(rows),
        "label": "+Fix1+2+3+4: IL from dashboard",
        "details": fix4_details,
    }

    # Fix5: Exclude negatives
    fix5_pos = sum(d["overpay_new"] for d in fix4_details if d["overpay_new"] >= 0)
    fix5_neg_count = sum(1 for d in fix4_details if d["overpay_new"] < 0)
    results["fix5"] = {
        "total_overpay": round(fix5_pos, 2),
        "rows_count": sum(1 for d in fix4_details if d["overpay_new"] >= 0),
        "rows_excluded": fix5_neg_count,
        "label": "+Fix1+2+3+4+5: Exclude negatives",
        "details": [d for d in fix4_details if d["overpay_new"] >= 0],
    }

    return results


# ─── Output ──────────────────────────────────────────────────────────────


def print_decomposition(results: dict[str, dict], reference: float | None = None):
    """Print decomposition table."""
    steps = ["baseline", "fix1", "fix2", "fix3", "fix4", "fix5"]
    labels = {
        "baseline": "Baseline (original audit)",
        "fix1": "Fix 1: Reverse logistics filter",
        "fix2": "Fix 2: Period-based tariffs",
        "fix3": "Fix 3: Warehouse coefficients",
        "fix4": "Fix 4: IL from WB dashboard",
        "fix5": "Fix 5: Exclude negatives",
    }

    logger.info(f"\n{'=' * 70}")
    logger.info("DISCREPANCY DECOMPOSITION REPORT")
    logger.info(f"{'=' * 70}")

    logger.info(f"\n{'Step':<45} {'Overpay':>14} {'Delta':>14} {'Rows':>8}")
    logger.info(f"{'-' * 82}")

    prev_total = None
    for step in steps:
        r = results[step]
        total = r["total_overpay"]
        delta = total - prev_total if prev_total is not None else 0
        delta_str = f"{delta:>+14,.2f}" if prev_total is not None else f"{'—':>14}"
        logger.info(f"{labels[step]:<45} {total:>14,.2f} {delta_str} {r['rows_count']:>8,}")
        prev_total = total

    logger.info(f"{'-' * 82}")

    baseline = results["baseline"]["total_overpay"]
    final = results["fix5"]["total_overpay"]
    total_correction = final - baseline
    logger.info(f"{'Total correction':<45} {'':>14} {total_correction:>+14,.2f}")
    logger.info(f"{'Final overpayment':<45} {final:>14,.2f}")

    if reference is not None:
        remainder = final - reference
        logger.info(f"{'Reference':<45} {reference:>14,.2f}")
        logger.info(f"{'Unexplained remainder':<45} {remainder:>+14,.2f}")
        logger.info(f"{'=' * 70}")
        if abs(remainder) < 1.0:
            logger.info("✓ TARGET MET: remainder < 1 rub")
        else:
            logger.info(f"Remainder: {remainder:,.2f} rub")
    else:
        logger.info(f"{'=' * 70}")

    # Per-fix impact
    logger.info(f"\n{'Fix':<45} {'Impact (rub)':>14}")
    logger.info(f"{'-' * 60}")
    prev = results["baseline"]["total_overpay"]
    for step in steps[1:]:
        r = results[step]
        delta = r["total_overpay"] - prev
        logger.info(f"{labels[step]:<45} {delta:>+14,.2f}")
        prev = r["total_overpay"]

    # Coef sources
    if "coef_sources" in results.get("fix3", {}):
        cs = results["fix3"]["coef_sources"]
        logger.info(f"\nCoefficient sources (Fix 3):")
        for src, cnt in cs.items():
            logger.info(f"  {src}: {cnt:,}")


def generate_final_excel(
    results: dict[str, dict],
    il_table: dict[date, float],
    tariffs_index: dict[str, list[tuple[date, float]]],
    volumes: dict[int, float],
    bonus_type_stats: dict[str, dict],
    output_path: Path,
):
    """Generate final 8-sheet Excel matching financiers' column format.

    Sheets:
      1. СВОД — summary by report_id
      2. Расчёт логистики (короб) — all forward rows with intermediate values
      3. Переплата по логистике — only positive overpay rows
      4. Переплата по артикулам — pivot by nm_id
      5. ИЛ — IL values by week
      6. Виды логистики — stats by bonus_type_name
      7. Габариты в карточке — nm_id → volume
      8. Тарифы короб — warehouse coefficients
    """
    logger.info(f"\nGenerating final Excel: {output_path.name}")

    fix4_details = results["fix4"].get("details", [])
    fix5_details = results["fix5"].get("details", [])

    def _fmt_date(d):
        if isinstance(d, date):
            return d.isoformat()
        return d

    # --- Helper: build row for Расчёт логистики columns ---
    def _logistics_row(d: dict) -> dict:
        return {
            "Номер отчета": d.get("report_id"),
            "Номер поставки": d.get("gi_id"),
            "Код номенклатуры": d["nm_id"],
            "Дата заказа покупателем": _fmt_date(d.get("order_dt")),
            "Услуги по доставке товара покупателю": d["delivery_rub"],
            "Дата начала действия фиксации": _fmt_date(d.get("fix_from")),
            "Дата конца действия фиксации": _fmt_date(d.get("fix_to")),
            "Склад": d.get("warehouse", ""),
            "ШК": d.get("shk_id"),
            "SRID": d.get("srid"),
            "Фиксированный коэффициент склада по поставке": d.get("fixed_coef", 0),
            "Коэф склада для расчета": d.get("coef_new", 0),
            "Объем товара из карточки товара": d.get("volume", 0),
            "КТР": d.get("il_new", 1.0),
            "Стоимость 1 л": d.get("base_1l", 0),
            "Стоимость доп.л": d.get("extra_l", 0),
            "Стоимость логистики (руб)": d.get("calc_cost_new", 0),
            "Разница": d.get("overpay_new", 0),
        }

    def _overpay_row(d: dict) -> dict:
        return {
            "Номер отчета": d.get("report_id"),
            "Номер поставки": d.get("gi_id"),
            "Код номенклатуры": d["nm_id"],
            "Дата заказа покупателем": _fmt_date(d.get("order_dt")),
            "Услуги по доставке товара покупателю": d["delivery_rub"],
            "Дата начала действия фиксации": _fmt_date(d.get("fix_from")),
            "Дата конца действия фиксации": _fmt_date(d.get("fix_to")),
            "Склад": d.get("warehouse", ""),
            "ШК": d.get("shk_id"),
            "SRID": d.get("srid"),
            "Фиксированный коэффициент склада по поставке": d.get("fixed_coef", 0),
            "Коэф склада для расчета": d.get("coef_new", 0),
            "Объем товара из карточки товара": d.get("volume", 0),
            "КТР": d.get("il_new", 1.0),
            "Стоимость 1 л": d.get("base_1l", 0),
            "Стоимость логистики (руб)": d.get("calc_cost_new", 0),
            "Сумма переплаты": d.get("overpay_new", 0),
        }

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # Sheet 1: СВОД — summary by report_id
        svod: dict[str, dict] = defaultdict(lambda: {"delivery_rub": 0.0, "overpay": 0.0, "count": 0})
        for d in fix5_details:
            rid = d.get("report_id") or "N/A"
            svod[rid]["delivery_rub"] += d["delivery_rub"]
            svod[rid]["overpay"] += d["overpay_new"]
            svod[rid]["count"] += 1
        svod_rows = []
        for rid in sorted(svod.keys(), key=lambda x: str(x)):
            s = svod[rid]
            svod_rows.append({
                "Номер отчета": rid,
                "Строк": s["count"],
                "Услуги по доставке (руб)": round(s["delivery_rub"], 2),
                "Сумма переплаты (руб)": round(s["overpay"], 2),
            })
        # Totals row
        if svod_rows:
            svod_rows.append({
                "Номер отчета": "ИТОГО",
                "Строк": sum(r["Строк"] for r in svod_rows),
                "Услуги по доставке (руб)": round(sum(r["Услуги по доставке (руб)"] for r in svod_rows), 2),
                "Сумма переплаты (руб)": round(sum(r["Сумма переплаты (руб)"] for r in svod_rows), 2),
            })
        pd.DataFrame(svod_rows).to_excel(writer, sheet_name="СВОД", index=False)

        # Sheet 2: Расчёт логистики (короб) — all forward rows
        if fix4_details:
            rows_calc = [_logistics_row(d) for d in fix4_details]
            pd.DataFrame(rows_calc).to_excel(writer, sheet_name="Расчёт логистики (короб)", index=False)

        # Sheet 3: Переплата по логистике — only positive overpay
        if fix5_details:
            rows_overpay = [_overpay_row(d) for d in fix5_details]
            pd.DataFrame(rows_overpay).to_excel(writer, sheet_name="Переплата по логистике", index=False)

        # Sheet 4: Переплата по артикулам — pivot by nm_id
        art_agg: dict[int, dict] = defaultdict(lambda: {"delivery_rub": 0.0, "overpay": 0.0, "count": 0})
        for d in fix5_details:
            nm = d["nm_id"]
            art_agg[nm]["delivery_rub"] += d["delivery_rub"]
            art_agg[nm]["overpay"] += d["overpay_new"]
            art_agg[nm]["count"] += 1
        art_rows = []
        for nm in sorted(art_agg.keys()):
            a = art_agg[nm]
            art_rows.append({
                "Код номенклатуры": nm,
                "Строк": a["count"],
                "Услуги по доставке (руб)": round(a["delivery_rub"], 2),
                "Сумма переплаты (руб)": round(a["overpay"], 2),
            })
        if art_rows:
            art_rows.append({
                "Код номенклатуры": "ИТОГО",
                "Строк": sum(r["Строк"] for r in art_rows),
                "Услуги по доставке (руб)": round(sum(r["Услуги по доставке (руб)"] for r in art_rows), 2),
                "Сумма переплаты (руб)": round(sum(r["Сумма переплаты (руб)"] for r in art_rows), 2),
            })
        pd.DataFrame(art_rows).to_excel(writer, sheet_name="Переплата по артикулам", index=False)

        # Sheet 5: ИЛ — IL values by week
        il_rows = []
        for mon in sorted(il_table.keys()):
            sun = mon + timedelta(days=6)
            il_rows.append({
                "Дата обновления": mon.isoformat(),
                "Неделя": f"{mon.isoformat()} — {sun.isoformat()}",
                "ИЛ": il_table[mon],
            })
        pd.DataFrame(il_rows).to_excel(writer, sheet_name="ИЛ", index=False)

        # Sheet 6: Виды логистики — stats by bonus_type_name
        bt_rows = []
        for bt_name, stats in sorted(bonus_type_stats.items()):
            bt_rows.append({
                "Вид логистики": bt_name,
                "Строк": stats.get("count", 0),
                "Услуги по доставке (руб)": round(stats.get("delivery_rub", 0), 2),
                "В аудите (прямая)": "Да" if bt_name in FORWARD_DELIVERY_TYPES else "Нет",
            })
        pd.DataFrame(bt_rows).to_excel(writer, sheet_name="Виды логистики", index=False)

        # Sheet 7: Габариты в карточке — nm_id → volume
        vol_rows = []
        for nm in sorted(volumes.keys()):
            vol_rows.append({
                "Код номенклатуры": nm,
                "Объем (л)": volumes[nm],
            })
        pd.DataFrame(vol_rows).to_excel(writer, sheet_name="Габариты в карточке", index=False)

        # Sheet 8: Тарифы короб — warehouse coefficients
        tariff_rows = []
        for wh in sorted(tariffs_index.keys()):
            for dt, coef in sorted(tariffs_index[wh]):
                tariff_rows.append({
                    "Дата": dt.isoformat(),
                    "Склад": wh,
                    "Коэффициент (%)": round(coef * 100, 2),
                })
        pd.DataFrame(tariff_rows).to_excel(writer, sheet_name="Тарифы короб", index=False)

    sheet_count = 8
    logger.info(f"  Saved: {output_path} ({sheet_count} sheets)")
    logger.info(f"  Расчёт логистики: {len(fix4_details):,} rows")
    logger.info(f"  Переплата: {len(fix5_details):,} rows, total: {sum(d['overpay_new'] for d in fix5_details):,.2f} rub")


# ─── Main ────────────────────────────────────────────────────────────────


def run_ooo():
    """Run OOO recalculation from original audit file."""
    logger.info("=" * 70)
    logger.info("OOO LOGISTICS AUDIT RECALCULATION")
    logger.info("=" * 70)

    tariffs_index = load_tariff_file()
    rows = load_ooo_data()

    results = recalculate_ooo(rows, tariffs_index, OOO_IL_DASHBOARD)
    print_decomposition(results)

    # Collect volumes dict from all rows (nm_id → volume)
    volumes: dict[int, float] = {}
    for r in rows:
        nm = r.get("nm_id")
        vol = r.get("volume", 0)
        if nm and vol > 0:
            volumes[nm] = vol

    # Collect bonus_type stats from all rows
    bonus_type_stats: dict[str, dict] = {}
    for r in rows:
        bt = r.get("bonus_type", "Unknown")
        if bt not in bonus_type_stats:
            bonus_type_stats[bt] = {"count": 0, "delivery_rub": 0.0}
        bonus_type_stats[bt]["count"] += 1
        bonus_type_stats[bt]["delivery_rub"] += r.get("delivery_rub", 0)

    output_path = SERVICE_DIR / "ООО Wookiee — Перерасчёт логистики (v2-final).xlsx"
    generate_final_excel(
        results=results,
        il_table=OOO_IL_DASHBOARD,
        tariffs_index=tariffs_index,
        volumes=volumes,
        bonus_type_stats=bonus_type_stats,
        output_path=output_path,
    )

    return results


def run_fisanov():
    """Validate algorithm on Fisanov reference."""
    logger.info("=" * 70)
    logger.info("FISANOV VALIDATION (reference: 144,901 rub)")
    logger.info("=" * 70)

    tariffs_index = load_tariff_file()
    rows = load_fisanov_rows()

    results = recalculate_fisanov(rows, tariffs_index, FISANOV_IL)
    print_decomposition(results, reference=144_901.0)

    return results


def main():
    parser = argparse.ArgumentParser(description="Recalculate OOO logistics audit")
    parser.add_argument("--fisanov", action="store_true", help="Validate on Fisanov")
    args = parser.parse_args()

    if args.fisanov:
        run_fisanov()
    else:
        run_ooo()


if __name__ == "__main__":
    main()
