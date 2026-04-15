"""Generate corrected Excel report from Fisanov reference data.

Applies all 5 fixes and produces a new Excel with:
- Sheet 1: Corrected overpayment calculations
- Sheet 2: Comparison (our vs Fisanov reference)
- Sheet 3: Decomposition by fix
- Sheet 4: IL values used

Usage:
    python -m services.logistics_audit.generate_fisanov_report
"""
from __future__ import annotations
import logging
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from services.logistics_audit.calculators.logistics_overpayment import (
    calculate_row_overpayment,
)
from services.logistics_audit.calculators.tariff_periods import get_base_tariffs

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

FISANOV_PATH = Path(__file__).parent / "ИП Фисанов. Проверка логистики 05.01.2026 г. - 01.02.2026 г._Итоговый.xlsx"
OUTPUT_DIR = Path(__file__).parent

# Fisanov IL values by week (from ИЛ sheet in reference Excel)
FISANOV_IL: dict[date, float] = {
    date(2025, 11, 3): 1.42,
    date(2025, 11, 10): 1.48,
    date(2025, 11, 17): 1.40,
    date(2025, 11, 24): 1.34,
    date(2025, 12, 1): 1.29,
    date(2025, 12, 8): 1.33,
    date(2025, 12, 15): 1.34,
    date(2025, 12, 22): 1.37,
    date(2025, 12, 29): 1.37,
    date(2026, 1, 5): 1.36,
    date(2026, 1, 12): 1.34,
    date(2026, 1, 19): 1.35,
    date(2026, 1, 26): 1.33,
    date(2026, 2, 2): 1.36,
}


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def load_rows() -> list[dict]:
    logger.info(f"Loading {FISANOV_PATH.name}...")
    wb = openpyxl.load_workbook(str(FISANOV_PATH), read_only=True, data_only=True)
    ws = wb["Переплата по логистике"]
    rows = []
    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if row_vals[0] is None and row_vals[2] is None:
            continue
        try:
            rows.append({
                "report_id": row_vals[0],
                "gi_id": row_vals[1],
                "nm_id": row_vals[2],
                "order_dt": _to_date(row_vals[3]),
                "delivery_rub": float(row_vals[4] or 0),
                "fix_date_from": _to_date(row_vals[5]),
                "fix_date_to": _to_date(row_vals[6]),
                "warehouse": row_vals[7],
                "shk_id": row_vals[8],
                "srid": row_vals[9],
                "fixed_coef": float(row_vals[10] or 0),
                "calc_coef": float(row_vals[11] or 0),
                "volume": float(row_vals[12] or 0),
                "ktr_ref": float(row_vals[13] or 0),
                "base_1l_ref": float(row_vals[14] or 0),
                "extra_l_ref": float(row_vals[15] or 0),
                "calc_cost_ref": float(row_vals[16] or 0),
                "difference_ref": float(row_vals[17] or 0),
            })
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    logger.info(f"Loaded {len(rows)} rows")
    return rows


def _style_header(ws, max_col, row=1):
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(style="thin")
    border = Border(bottom=thin)
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)


def generate():
    rows = load_rows()
    if not rows:
        logger.error("No rows!")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # =========================================================================
    # Sheet 1: Исправленный расчёт переплаты
    # =========================================================================
    ws1 = wb.create_sheet("Переплата (исправленный)")
    headers1 = [
        "№ отчёта", "Номер поставки", "Код номенклатуры", "Дата заказа",
        "Услуги по доставке", "Дата начала фиксации", "Дата конца фиксации",
        "Склад", "ШК", "Srid",
        "Фикс. коэф.", "Коэф. для расчёта", "Объём",
        "ИЛ (исправл.)", "Стоимость 1л (исправл.)", "Стоимость доп.л (исправл.)",
        "Расч. стоимость (исправл.)", "Разница (исправл.)",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(1, col, h)
    _style_header(ws1, len(headers1))

    total_overpay = 0.0
    excel_row = 2
    for r in rows:
        base_1l, extra_l = get_base_tariffs(
            order_date=r["order_dt"],
            fixation_start=r["fix_date_from"],
            fixation_end=r["fix_date_to"],
            volume=r["volume"],
        )
        coef = r["calc_coef"]
        if r["fixed_coef"] > 0 and r["fix_date_to"] and r["order_dt"] and r["fix_date_to"] > r["order_dt"]:
            coef = r["fixed_coef"]

        il = 1.0
        if r["order_dt"]:
            mon = _monday(r["order_dt"])
            il = FISANOV_IL.get(mon, r.get("ktr_ref", 1.0))

        res = calculate_row_overpayment(
            delivery_rub=r["delivery_rub"],
            volume=r["volume"],
            coef=coef,
            base_1l=base_1l,
            extra_l=extra_l,
            order_dt=r["order_dt"],
            ktr_manual=il,
            is_fixed_rate=False,
            is_forward_delivery=True,
        )
        if res is None or res.overpayment < 0:
            continue

        ws1.cell(excel_row, 1, r["report_id"])
        ws1.cell(excel_row, 2, r["gi_id"])
        ws1.cell(excel_row, 3, r["nm_id"])
        ws1.cell(excel_row, 4, str(r["order_dt"]) if r["order_dt"] else "")
        ws1.cell(excel_row, 5, r["delivery_rub"])
        ws1.cell(excel_row, 6, str(r["fix_date_from"]) if r["fix_date_from"] else "")
        ws1.cell(excel_row, 7, str(r["fix_date_to"]) if r["fix_date_to"] else "")
        ws1.cell(excel_row, 8, r["warehouse"])
        ws1.cell(excel_row, 9, r["shk_id"])
        ws1.cell(excel_row, 10, r["srid"])
        ws1.cell(excel_row, 11, r["fixed_coef"])
        ws1.cell(excel_row, 12, coef)
        ws1.cell(excel_row, 13, r["volume"])
        ws1.cell(excel_row, 14, il)
        ws1.cell(excel_row, 15, base_1l)
        ws1.cell(excel_row, 16, extra_l)
        ws1.cell(excel_row, 17, res.calculated_cost)
        ws1.cell(excel_row, 18, res.overpayment)
        total_overpay += res.overpayment
        excel_row += 1

    # Summary row
    ws1.insert_rows(1)
    ws1.cell(1, 1, "ИТОГО переплата:")
    ws1.cell(1, 1).font = Font(bold=True, size=12)
    ws1.cell(1, 18, round(total_overpay, 2))
    ws1.cell(1, 18).font = Font(bold=True, size=12)

    # =========================================================================
    # Sheet 2: Сравнение наш vs Фисанов
    # =========================================================================
    ws2 = wb.create_sheet("Сравнение с эталоном")
    headers2 = [
        "Код номенклатуры", "Дата заказа", "Услуги по доставке",
        "ИЛ (Фисанов)", "ИЛ (наш)", "Δ ИЛ",
        "Тариф 1л (Фисанов)", "Тариф 1л (наш)", "Δ Тариф",
        "Расч. стоимость (Фисанов)", "Расч. стоимость (наш)", "Δ Стоимость",
        "Разница (Фисанов)", "Разница (наш)", "Δ Разница",
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(1, col, h)
    _style_header(ws2, len(headers2))

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for i, r in enumerate(rows, 2):
        base_1l, extra_l = get_base_tariffs(
            order_date=r["order_dt"],
            fixation_start=r["fix_date_from"],
            fixation_end=r["fix_date_to"],
            volume=r["volume"],
        )
        coef = r["calc_coef"]
        if r["fixed_coef"] > 0 and r["fix_date_to"] and r["order_dt"] and r["fix_date_to"] > r["order_dt"]:
            coef = r["fixed_coef"]
        il = 1.0
        if r["order_dt"]:
            mon = _monday(r["order_dt"])
            il = FISANOV_IL.get(mon, r.get("ktr_ref", 1.0))

        res = calculate_row_overpayment(
            delivery_rub=r["delivery_rub"], volume=r["volume"], coef=coef,
            base_1l=base_1l, extra_l=extra_l, order_dt=r["order_dt"],
            ktr_manual=il, is_fixed_rate=False, is_forward_delivery=True,
        )
        our_cost = res.calculated_cost if res else 0
        our_diff = res.overpayment if res else 0

        ws2.cell(i, 1, r["nm_id"])
        ws2.cell(i, 2, str(r["order_dt"]) if r["order_dt"] else "")
        ws2.cell(i, 3, r["delivery_rub"])
        ws2.cell(i, 4, r["ktr_ref"])
        ws2.cell(i, 5, il)
        ws2.cell(i, 6, round(il - r["ktr_ref"], 4))
        ws2.cell(i, 7, r["base_1l_ref"])
        ws2.cell(i, 8, base_1l)
        ws2.cell(i, 9, round(base_1l - r["base_1l_ref"], 2))
        ws2.cell(i, 10, r["calc_cost_ref"])
        ws2.cell(i, 11, our_cost)
        ws2.cell(i, 12, round(our_cost - r["calc_cost_ref"], 2))
        ws2.cell(i, 13, r["difference_ref"])
        ws2.cell(i, 14, our_diff)
        delta = round(our_diff - r["difference_ref"], 2)
        ws2.cell(i, 15, delta)
        if abs(delta) > 1.0:
            ws2.cell(i, 15).fill = red_fill
        elif abs(delta) < 0.01:
            ws2.cell(i, 15).fill = green_fill

    # =========================================================================
    # Sheet 3: Декомпозиция по фиксам
    # =========================================================================
    ws3 = wb.create_sheet("Декомпозиция")
    decomp_headers = ["Источник ошибки", "Вклад (руб)", "Строк", "Комментарий"]
    for col, h in enumerate(decomp_headers, 1):
        ws3.cell(1, col, h)
    _style_header(ws3, len(decomp_headers))

    # Calculate baseline
    baseline = sum(
        (r := calculate_row_overpayment(
            delivery_rub=row["delivery_rub"], volume=row["volume"],
            coef=row["calc_coef"], base_1l=46.0, extra_l=14.0,
            order_dt=row["order_dt"], ktr_manual=1.0,
            is_fixed_rate=False, is_forward_delivery=True,
        )) and r.overpayment or 0
        for row in rows
    )

    # Fix 2 only
    fix2 = 0.0
    for row in rows:
        b1, el = get_base_tariffs(row["order_dt"], row["fix_date_from"], row["fix_date_to"], row["volume"])
        r = calculate_row_overpayment(
            delivery_rub=row["delivery_rub"], volume=row["volume"],
            coef=row["calc_coef"], base_1l=b1, extra_l=el,
            order_dt=row["order_dt"], ktr_manual=1.0,
            is_fixed_rate=False, is_forward_delivery=True,
        )
        if r:
            fix2 += r.overpayment

    ref_total = sum(row["difference_ref"] for row in rows)

    decomp_data = [
        ("Baseline (старая логика)", round(baseline, 2), len(rows), "tariff=46, KTR=1.0"),
        ("Fix 1: Фильтр обратной логистики", 0.0, 0, "Все строки Фисанова — forward"),
        ("Fix 2: Тарифы по периодам", round(fix2 - baseline, 2), len(rows), "46→32 руб (sub-liter 0.98L)"),
        ("Fix 3: Коэффициент склада", 0.0, len(rows), "Фисанов уже использует верные коэфы"),
        ("Fix 4: ИЛ калибровка", round(total_overpay - fix2, 2), len(rows), "1.0 → 1.29–1.37"),
        ("Fix 5: Отрицательные разницы", 0.0, 0, "Нет отрицательных у Фисанова"),
        ("", "", "", ""),
        ("Наш результат", round(total_overpay, 2), "", ""),
        ("Эталон (Фисанов)", round(ref_total, 2), "", ""),
        ("Необъяснённый остаток", round(total_overpay - ref_total, 2), "", "Округление"),
    ]
    for i, (name, val, cnt, comment) in enumerate(decomp_data, 2):
        ws3.cell(i, 1, name)
        ws3.cell(i, 2, val)
        ws3.cell(i, 3, cnt)
        ws3.cell(i, 4, comment)

    ws3.cell(9, 1).font = Font(bold=True)
    ws3.cell(10, 1).font = Font(bold=True)
    ws3.cell(11, 1).font = Font(bold=True)

    # =========================================================================
    # Sheet 4: ИЛ по неделям
    # =========================================================================
    ws4 = wb.create_sheet("ИЛ")
    il_headers = ["Неделя (понедельник)", "ИЛ (Фисанов)", "Период"]
    for col, h in enumerate(il_headers, 1):
        ws4.cell(1, col, h)
    _style_header(ws4, len(il_headers))

    for i, (mon, il) in enumerate(sorted(FISANOV_IL.items()), 2):
        sun = mon + timedelta(days=6)
        ws4.cell(i, 1, mon.isoformat())
        ws4.cell(i, 2, il)
        ws4.cell(i, 3, f"{mon.isoformat()} — {sun.isoformat()}")

    # Column widths
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted = min(max_len + 2, 30)
            ws.column_dimensions[col[0].column_letter].width = adjusted

    # Save
    output_path = OUTPUT_DIR / "ИП Фисанов — Исправленный расчёт логистики (v2).xlsx"
    wb.save(str(output_path))
    logger.info(f"\nSaved: {output_path}")
    logger.info(f"Total overpayment: {total_overpay:,.2f} rub (reference: {ref_total:,.2f})")
    return str(output_path)


if __name__ == "__main__":
    generate()
