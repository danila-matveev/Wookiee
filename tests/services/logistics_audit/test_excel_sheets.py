import openpyxl
from io import BytesIO
from services.logistics_audit.output.sheet_overpayment_formulas import write_overpayment_formulas
from services.logistics_audit.models.report_row import ReportRow


def _make_row(**overrides) -> ReportRow:
    defaults = dict(
        realizationreport_id=658038623, nm_id=257131227,
        office_name="Коледино", supplier_oper_name="Логистика",
        bonus_type_name="К клиенту при продаже", delivery_rub=147.23,
        dlv_prc=1.95, fix_tariff_date_from=None, fix_tariff_date_to=None,
        order_dt=None, shk_id=41793344171, srid="20966880121115600.0.0",
        gi_id=12345, gi_box_type_name="Микс", storage_fee=0, penalty=0,
        deduction=0, rebill_logistic_cost=0, ppvz_for_pay=0,
        ppvz_supplier_name="", retail_amount=0, date_from="", date_to="",
        doc_type_name="", acceptance=0,
    )
    defaults.update(overrides)
    return ReportRow(**defaults)


def test_sheet_overpayment_formulas_header():
    """Sheet 1 has correct header columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [_make_row()]
    write_overpayment_formulas(ws, rows, ktr=1.04, base_1l=46.0, extra_l=14.0)
    headers = [ws.cell(1, c).value for c in range(1, 20)]
    assert "Код номенклатуры" in headers
    assert "Услуги по доставке" in headers
    assert "Разница" in headers


def test_sheet_overpayment_formulas_has_formula():
    """Column R (стоимость логистики) must contain an Excel formula."""
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [_make_row()]
    write_overpayment_formulas(ws, rows, ktr=1.04, base_1l=46.0, extra_l=14.0)
    # Row 2 is first data row
    formula_cell = ws.cell(2, 18).value  # R = column 18
    assert formula_cell is not None
    assert str(formula_cell).startswith("=")


def test_generate_full_workbook():
    """Excel generator creates workbook with all 11 sheet names."""
    from services.logistics_audit.output.excel_generator import generate_workbook
    from services.logistics_audit.models.audit_config import AuditConfig
    from datetime import date

    row = _make_row(raw={"realizationreport_id": 658038623, "nm_id": 257131227})
    config = AuditConfig(
        api_key="test", date_from=date(2026, 3, 9), date_to=date(2026, 3, 15),
        ktr=1.04, base_tariff_1l=46.0, base_tariff_extra_l=14.0,
    )
    wb = generate_workbook(
        config=config,
        all_rows=[row],
        logistics_rows=[row],
        overpayment_results=[None],
        coefs=[1.95],
        card_dims={257131227: {"width": 33, "height": 22, "length": 4, "volume": 2.904}},
        tariffs_box={},
        tariffs_pallet={},
        wb_volumes={},
    )
    sheet_names = wb.sheetnames
    assert len(sheet_names) == 11
    assert "Переплата по логистике (короб)" in sheet_names
    assert "СВОД" in sheet_names
    assert "Детализация" in sheet_names
